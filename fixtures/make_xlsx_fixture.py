"""One-off script to generate the sample questionnaire fixture. Run once after deps are installed."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "Questionnaire"

headers = ["Section", "Question", "Vendor Response", "Compliance Status"]
for col, text in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=text)
    cell.font = Font(bold=True)

rows = [
    ("section", "Section 1: Access Control"),
    ("question", "Do you require multi-factor authentication for all employee access to production systems?"),
    ("blank", None),
    ("question", "Is all data encrypted in transit between clients and your production services?"),
    ("question", "Do you encrypt data at rest and in transit, and how often are encryption keys rotated?"),
    ("section", "Section 2: Business Continuity"),
    ("question", "How frequently are production databases backed up and what is your backup retention period?"),
    ("question", "What is your disaster recovery RTO and RPO?"),
    ("question", "Do you maintain a public bug bounty program?"),
]

row_num = 2
for kind, text in rows:
    if kind == "section":
        ws.cell(row=row_num, column=1, value=text).font = Font(bold=True)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    elif kind == "question":
        ws.cell(row=row_num, column=2, value=text)
    # "blank" rows: leave the whole row empty on purpose
    row_num += 1

last_row = row_num - 1
dv = DataValidation(type="list", formula1='"Yes,No,Not Applicable,Compensating Control"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"D3:D{last_row}")

for col, width in zip("ABCD", (24, 60, 40, 22)):
    ws.column_dimensions[col].width = width

out_path = Path(__file__).parent / "questionnaire_sample.xlsx"
wb.save(str(out_path))
print(f"Wrote {out_path}")
