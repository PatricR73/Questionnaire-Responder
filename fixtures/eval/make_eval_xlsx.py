"""One-off script to generate an .xlsx questionnaire from the eval fixture's 20
questions, so the eval set can be run through the real pipeline CLI
(`python -m src.pipeline answer ...`) exactly the way a real questionnaire is —
same code path as any other run, not a parallel implementation. Run this only if
questions.json changes; the generated file is committed alongside it.

Reads question text and order directly from questions.json rather than duplicating
it here, so the two can never silently drift apart.

Question text is quoted verbatim from CAIQ v4.0.2 — see LABELING_GUIDE.md's
licensing section. This workbook reproduces the same small, attributed selection,
in a different file format, not additional content.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font

HERE = Path(__file__).parent
with open(HERE / "questions.json") as f:
    questions = json.load(f)["questions"]

wb = Workbook()
ws = wb.active
ws.title = "Eval Questions"

wb.properties.title = "CAIQ v4.0.2 eval fixture — 20 questions"
wb.properties.description = (
    "Question text quoted verbatim from the Cloud Security Alliance CAIQ v4.0.2 "
    "under fair use with attribution — see fixtures/eval/LABELING_GUIDE.md. "
    "Generated from questions.json by make_eval_xlsx.py; do not hand-edit."
)
ws["A1"].comment = Comment(
    "Generated from fixtures/eval/questions.json by make_eval_xlsx.py. Edit that "
    "file and regenerate, not this one directly.",
    "Questionnaire Responder",
)

headers = ["Source ID", "Question", "Answer"]
for col, text in enumerate(headers, start=1):
    ws.cell(row=1, column=col, value=text).font = Font(bold=True)

for i, q in enumerate(questions, start=2):
    ws.cell(row=i, column=1, value=q["source_id"])
    ws.cell(row=i, column=2, value=q["question_text"])

for col, width in zip("ABC", (14, 90, 60)):
    ws.column_dimensions[col].width = width

out_path = HERE / "questionnaire_eval.xlsx"
wb.save(str(out_path))
print(f"Wrote {out_path} ({len(questions)} rows)")
