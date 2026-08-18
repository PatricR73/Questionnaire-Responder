"""Reproducible eval runner: runs the real pipeline CLI against the 20-question CAIQ
fixture and scores the result against fixtures/eval/questions.json.

Deliberately shells out to `python -m src.pipeline answer` rather than calling
generate_answer/AnthropicAnswerer directly — this project's headline numbers were
first produced by one-off scripts that bypassed the CLI, which meant the thing being
measured wasn't quite the thing anyone would actually run. This script IS the
documented command; the CLI path and the eval path are the same code.

Usage (from the project root, with ANTHROPIC_API_KEY exported):

    python fixtures/eval/run_eval.py
    python fixtures/eval/run_eval.py --repeats 3

Prints, per question: source_id, expected_label, actual status/confidence/polarity,
the full drafted answer text, and a structural match flag. Then a summary, with the
6 NOT_FOUND questions called out on their own line — a NOT_FOUND question that comes
back ANSWERED is flagged as a regression regardless of the rest of the score, per
this project's tuning-log convention.

--repeats N runs the whole pipeline N times into distinct output paths
(out/eval_run_1.xlsx .. out/eval_run_N.xlsx; plain out/eval_run.xlsx for N=1) so a
reported delta between runs is never a single stochastic sample again: every question
gets a per-question stability figure (how many of the N runs produced the same
status), the summary prints min/mean/max structural match across runs, and the 95%
Wilson confidence interval for the structural-match proportion at the current eval-set
size — with n=20, a one-question delta is squarely inside the interval, which is the
whole point of printing it.

This script does NOT assign usable/needs-editing/wrong — that's a hand-scored
judgment call by design (see LABELING_GUIDE.md's anti-mirror methodology), not
something to automate. It gives you the structural match plus every full answer to
read, which is the input hand-scoring needs.
"""

import argparse
import json
import statistics
import subprocess
import sys
from collections import Counter
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
QUESTIONNAIRE = REPO_ROOT / "fixtures" / "eval" / "questionnaire_eval.xlsx"
QUESTIONS = REPO_ROOT / "fixtures" / "eval" / "questions.json"
OUTPUT = REPO_ROOT / "out" / "eval_run.xlsx"

sys.path.insert(0, str(REPO_ROOT))

from src.store.db import connect

EXPECTED_STATUS = {
    "ANSWERED_AFFIRMS": "ANSWERED",
    "ANSWERED_DENIES": "ANSWERED",
    "ANSWERED_PARTIAL": "ANSWERED",
    "AMBIGUOUS_EVIDENCE": "ANSWERED",
    "NOT_FOUND": "NOT_FOUND",
}
# Expected polarity for each label, compared for every ANSWERED row. A label that
# expects a polarity and gets a different one is a structural mismatch — the
# harness used to print polarity without ever comparing it, so a question labeled
# ANSWERED_DENIES that came back affirming scored as a match, blind to exactly the
# distinction (we prohibit shared accounts vs we permit them) the polarity field
# was introduced to capture.
#
# AMBIGUOUS_EVIDENCE maps to "partial", per system-prompt rule 8: a contradiction
# must be surfaced as both claims attributed to their sources with polarity
# "partial" — never reconciled into one coherent polarity — so "partial" is the
# only polarity that counts as a match for those fixtures. This is a deliberate
# decision made here and recorded in TUNING_LOG.md; it matches how rule 8 already
# routes contradictions through the existing partial/low path.
EXPECTED_POLARITY = {
    "ANSWERED_AFFIRMS": "affirms",
    "ANSWERED_DENIES": "denies",
    "ANSWERED_PARTIAL": "partial",
    "AMBIGUOUS_EVIDENCE": "partial",
    "NOT_FOUND": None,
}
NOT_FOUND_LABEL = "NOT_FOUND"
# An affirms<->denies swap is a materially worse outcome than a missed answer — the
# row asserts the OPPOSITE of the documented fact — so it gets its own called-out
# regression line rather than being buried in the aggregate (same convention as the
# NOT_FOUND regression line).
_POLARITY_INVERSIONS = {"affirms": "denies", "denies": "affirms"}

# 1.959963984540054 = the 97.5th percentile of the standard normal (z for 95%).
_WILSON_Z = 1.959963984540054


def run_pipeline(output_path: Path, config_file: Path | None = None, provider: str = "anthropic") -> None:
    cmd = [
        sys.executable,
        "-m",
        "src.pipeline",
        "answer",
        "--questionnaire",
        str(QUESTIONNAIRE),
        "--output",
        str(output_path),
        "--limit",
        "0",
        "--provider",
        provider,
    ]
    if config_file is not None:
        cmd += ["--config", str(config_file)]
    print(f"Running: {' '.join(cmd)}")
    result = subprocess.run(cmd, cwd=REPO_ROOT, check=False)
    if result.returncode != 0:
        raise SystemExit(f"pipeline run failed (exit {result.returncode})")


def load_run_rows(output_path: Path) -> list[dict]:
    """Pull one run's scored rows out of the store, joined to the fixture labels."""
    source_id_by_row = {}
    import openpyxl

    ws = openpyxl.load_workbook(QUESTIONNAIRE).active
    for row in range(2, ws.max_row + 1):
        source_id = ws.cell(row=row, column=1).value
        if source_id:
            source_id_by_row[row] = source_id

    with open(QUESTIONS) as f:
        questions = {q["source_id"]: q for q in json.load(f)["questions"]}

    conn = connect(REPO_ROOT / "out" / "store.db")
    run_id = conn.execute(
        "SELECT id FROM questionnaire_runs WHERE output_path = ? ORDER BY id DESC LIMIT 1",
        (str(output_path),),
    ).fetchone()["id"]

    rows = conn.execute(
        """
        SELECT a.row_index, a.final_confidence, a.polarity, a.drafted_answer
        FROM answers a WHERE a.run_id = ? ORDER BY a.row_index
        """,
        (run_id,),
    ).fetchall()

    scored = []
    for row in rows:
        source_id = source_id_by_row.get(row["row_index"])
        q = questions.get(source_id)
        if q is None:
            continue
        expected_label = q["expected_label"]
        status = {"high": "ANSWERED", "low": "ANSWERED", "none": "NOT_FOUND", "error": "ERROR"}.get(
            row["final_confidence"], "ERROR"
        )
        scored.append(
            {
                "source_id": source_id,
                "row_index": row["row_index"],
                "expected_label": expected_label,
                "status": status,
                "confidence": row["final_confidence"],
                "polarity": row["polarity"],
                "answer": row["drafted_answer"],
            }
        )
    return scored


def structural_match(expected_label: str, status: str, polarity: str | None) -> bool:
    """Status/polarity vs. expected label. Polarity is compared for every ANSWERED
    label that expects one (see EXPECTED_POLARITY); NOT_FOUND rows match on status
    alone."""
    if status != EXPECTED_STATUS[expected_label]:
        return False
    expected_polarity = EXPECTED_POLARITY[expected_label]
    if expected_polarity is None:
        return True
    return polarity == expected_polarity


def score_run(rows: list[dict]) -> tuple[int, list[str], list[str]]:
    matches = 0
    not_found_regressions = []
    polarity_inversions = []
    for r in rows:
        expected_label = r["expected_label"]
        status = r["status"]
        polarity = r["polarity"]
        matches += structural_match(expected_label, status, polarity)
        if expected_label == NOT_FOUND_LABEL and status == "ANSWERED":
            not_found_regressions.append(r["source_id"])
        if (
            EXPECTED_POLARITY[expected_label] in _POLARITY_INVERSIONS
            and polarity == _POLARITY_INVERSIONS[EXPECTED_POLARITY[expected_label]]
        ):
            polarity_inversions.append(r["source_id"])
    return matches, not_found_regressions, polarity_inversions


def wilson_interval(k: int, n: int) -> tuple[float, float]:
    """95% Wilson score interval for a binomial proportion k/n — the honest uncertainty
    band around a structural-match count at this eval-set size. Chosen over the normal
    approximation because k/n near 0 or 1 (and small n) makes the normal interval
    degenerate; Wilson stays inside [0, 1] and is accurate even at n=20."""
    if n <= 0:
        return (0.0, 1.0)
    p = k / n
    z = _WILSON_Z
    denom = 1 + z * z / n
    center = (p + z * z / (2 * n)) / denom
    half = z * (p * (1 - p) / n + z * z / (4 * n * n)) ** 0.5 / denom
    return (max(0.0, center - half), min(1.0, center + half))


def print_run_detail(rows: list[dict], not_found_regressions: list[str], polarity_inversions: list[str]) -> None:
    print()
    print("=" * 100)
    for r in rows:
        is_match = structural_match(r["expected_label"], r["status"], r["polarity"])
        print(
            f"{r['source_id']:10s} expected={r['expected_label']:20s} status={r['status']:10s} "
            f"confidence={r['confidence']:6s} polarity={r['polarity']!s:8s} "
            f"match={'yes' if is_match else 'NO'}"
        )
        print(f"  answer: {r['answer'] or '(empty)'}")
        print()

    print("=" * 100)
    print("NOT_FOUND questions (must all stay NOT_FOUND):")
    for r in rows:
        if r["expected_label"] == NOT_FOUND_LABEL:
            print(f"  {r['source_id']}: {r['status']}")
    print()
    if not_found_regressions:
        print(f"!!! REGRESSION: these NOT_FOUND questions came back ANSWERED: {not_found_regressions}")
    else:
        print("No NOT_FOUND regressions.")
    if polarity_inversions:
        print(
            f"!!! POLARITY INVERSION: these questions answered the OPPOSITE of their expected polarity: {polarity_inversions}"
        )
    else:
        print("No polarity inversions.")


def main():
    # stdout is block-buffered when redirected to a file, so the parent's progress
    # lines interleave confusingly with the (OS-unbuffered) pipeline subprocess
    # output; force line buffering so a logged run reads in order.
    sys.stdout.reconfigure(line_buffering=True)

    parser = argparse.ArgumentParser(description=__doc__.split("Usage")[0])
    parser.add_argument(
        "--repeats",
        type=int,
        default=1,
        help="Run the pipeline N times into distinct output paths and report per-question "
        "stability plus min/mean/max structural match across runs (default 1).",
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=None,
        help="TOML file of tuning knobs passed through to the pipeline (see src/config.py) "
        "so a sweep is a loop over config files instead of a series of source edits.",
    )
    parser.add_argument(
        "--provider",
        type=str,
        default="openai-compatible",
        help="Generation provider: openai-compatible (default), anthropic, stub, or local (alias; warns on hosted URLs) — any OpenAI-compatible "
        "endpoint via QRESP_LOCAL_BASE_URL/QRESP_LOCAL_MODEL). The local provider lets the "
        "eval measure the fully-on-premise path (pack 3, C7) against the same scoring.",
    )
    args = parser.parse_args()
    repeats = args.repeats
    if repeats < 1:
        raise SystemExit("--repeats must be >= 1")

    if repeats == 1:
        outputs = [OUTPUT]
    else:
        outputs = [OUTPUT.with_stem(f"{OUTPUT.stem}_{i}") for i in range(1, repeats + 1)]

    per_run_rows: list[list[dict]] = []
    per_run_matches: list[int] = []
    per_run_regressions: list[list[str]] = []
    per_run_inversions: list[list[str]] = []

    for i, output_path in enumerate(outputs, start=1):
        print(f"\n=== Repeat {i}/{repeats} (output: {output_path.name}) ===")
        run_pipeline(output_path, config_file=args.config, provider=args.provider)
        rows = load_run_rows(output_path)
        matches, not_found_regressions, polarity_inversions = score_run(rows)
        per_run_rows.append(rows)
        per_run_matches.append(matches)
        per_run_regressions.append(not_found_regressions)
        per_run_inversions.append(polarity_inversions)
        print_run_detail(rows, not_found_regressions, polarity_inversions)

    n_questions = len(per_run_rows[0])

    if repeats > 1:
        print()
        print("=" * 100)
        print(f"Per-question stability across {repeats} runs (modal status / how many runs produced it):")
        by_source: dict[str, list[str]] = {}
        for rows in per_run_rows:
            for r in rows:
                by_source.setdefault(r["source_id"], []).append(r["status"])
        for source_id, statuses in by_source.items():
            counts = Counter(statuses)
            modal, modal_count = counts.most_common(1)[0]
            unstable = f"  <- statuses: {statuses}" if modal_count < len(statuses) else ""
            print(f"  {source_id:10s} {modal:10s} {modal_count}/{len(statuses)}{unstable}")

    print()
    print("=" * 100)
    print(f"Structural match per run: {per_run_matches}")
    print(f"NOT_FOUND regressions per run: {per_run_regressions}")
    print(f"Polarity inversions per run: {per_run_inversions}")
    mean_matches = statistics.mean(per_run_matches)
    print(
        f"min/mean/max across {repeats} run(s): {min(per_run_matches)} / {mean_matches:.1f} / "
        f"{max(per_run_matches)} out of {n_questions}"
    )

    # Wilson CI for the structural-match proportion at the current eval-set size. Uses
    # the mean match count across runs so a multi-run comparison has a single honest
    # band; at n=20 a one-question delta (e.g. 12/20 vs 13/20) sits entirely inside
    # it — which is exactly why the interval is printed: no delta under ~5 points at
    # this n is distinguishable from sampling noise.
    mean_k = round(mean_matches)
    ci_lo, ci_hi = wilson_interval(mean_k, n_questions)
    print(
        f"95% Wilson CI for structural-match proportion (n={n_questions}, mean across {repeats} run(s)): "
        f"[{ci_lo:.2f}, {ci_hi:.2f}]"
    )

    if repeats == 1:
        print()
        print("Structural match is NOT the usable/needs-editing/wrong score — read the full")
        print("answers above and hand-score them, per this project's methodology.")


if __name__ == "__main__":
    main()
