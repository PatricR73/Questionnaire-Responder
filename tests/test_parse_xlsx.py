"""P22: the column-detection heuristics in parse_xlsx, tested directly.

_find_column used to return the FIRST keyword match, and QUESTION_KEYWORDS includes
"control" — so a real CAIQ v4 sheet, which carries "Control ID" and "Control
Domain" columns before the question column, would pick the wrong column. Likewise
ANSWER_KEYWORDS includes "response", and real sheets carry multiple response
columns. Detection is now scoring-based: exact matches beat substrings, known
decoys are penalized, and the highest scorer wins.
"""

import openpyxl
import pytest

from src.questionnaire.parse_xlsx import detect_columns, iter_question_sheets, read_questions


def _sheet(headers, *, header_row=1, extra_rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, text in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=text)
    for row in extra_rows or []:
        for col, value in row.items():
            ws.cell(row=col if False else value[0], column=value[1], value=value[2])
    return ws


def _write_row(ws, row, values):
    for col, value in enumerate(values, start=1):
        ws.cell(row=row, column=col, value=value)


def test_caiq_shaped_sheet_picks_the_question_column_not_control_id():
    # Real CAIQ sheets put "Control ID" and "Control Domain" BEFORE the question
    # column; first-match detection grabbed "Control ID". Scoring must land on
    # the actual question column.
    ws = _sheet(["Control ID", "Control Domain", "Question", "Response"])
    cm = detect_columns(ws)
    assert cm.question_col == 3
    assert cm.answer_col == 4
    assert cm.vocab_col is None


def test_multiple_answer_like_columns_pick_the_exact_one():
    # Several columns contain "response"; the exact "Response" must win over
    # "Response Notes" even when it comes later.
    ws = _sheet(["Question", "Response Notes", "Response"])
    cm = detect_columns(ws)
    assert cm.answer_col == 3


def test_header_row_not_in_row_one():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Company Security Questionnaire"
    ws["A2"] = "Prepared for: customer"
    _write_row(ws, 3, ["Question", "Vendor Response"])
    cm = detect_columns(ws)
    assert cm.header_row == 3
    assert cm.question_col == 1
    assert cm.answer_col == 2


def test_merged_section_headers_and_blank_spacers_are_skipped():
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_row(ws, 1, ["Question", "Vendor Response"])
    ws.merge_cells("A2:B2")
    ws["A2"] = "Section 1: Access Control"  # merged section header — not a question
    _write_row(ws, 3, ["Q1?"])  # blank spacer row 4
    _write_row(ws, 5, ["Q2?"])
    _write_row(ws, 6, [None])  # blank spacer
    _write_row(ws, 7, ["Q3?"])
    questions = read_questions(ws, detect_columns(ws))
    assert [q.row_index for q in questions] == [3, 5, 7]
    assert [q.question_text for q in questions] == ["Q1?", "Q2?", "Q3?"]


def test_sheet_with_no_vocabulary_column():
    ws = _sheet(["Question", "Vendor Response"])
    cm = detect_columns(ws)
    assert cm.vocab_col is None
    assert cm.vocab_values is None


def test_sheet_with_vocabulary_column_and_dropdown():
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_row(ws, 1, ["Question", "Vendor Response", "Compliance Status"])
    from openpyxl.worksheet.datavalidation import DataValidation

    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv.add("C2:C100")
    ws.add_data_validation(dv)
    cm = detect_columns(ws)
    assert cm.vocab_col == 3
    assert cm.vocab_values == ["Yes", "No"]


def test_missing_header_row_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_row(ws, 1, ["Just some", "random content"])
    with pytest.raises(ValueError, match="header row"):
        detect_columns(ws)


def test_multi_sheet_workbook_with_instructions_first(tmp_path):
    """Pack 3, C6: the most likely first-contact failure — a workbook whose ACTIVE
    (first) tab is an instructions sheet and whose questions live on a later tab.
    workbook.active used to silently ignore everything but the first tab."""
    from openpyxl import Workbook

    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    instructions["A1"] = "How to use this questionnaire"
    instructions["A2"] = "Answer every row. Upload the completed file."
    qa = wb.create_sheet("Questionnaire")
    qa["A1"] = "Control ID"
    qa["B1"] = "Question"
    qa["C1"] = "Answer"
    qa["A2"] = "BCR-01.1"
    qa["B2"] = "Is cloud data periodically backed up?"
    qa["A3"] = "IAM-02.2"
    qa["B3"] = "Are access reviews automated?"
    path = tmp_path / "multi.xlsx"
    wb.save(path)

    workbook = openpyxl.load_workbook(path)
    assert workbook.active.title == "Instructions", "fixture must put instructions first"

    sheets = iter_question_sheets(workbook)
    assert [name for name, _, _ in sheets] == ["Questionnaire"]
    _, ws, column_map = sheets[0]
    questions = read_questions(ws, column_map)
    assert len(questions) == 2
    assert column_map.question_col == 2 and column_map.answer_col == 3


def test_column_override_map(tmp_path):
    """--map must override detection entirely: a sheet whose headers are
    unconventional is a dead end without the override."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Custom"
    ws["A1"] = "Control ID"
    ws["B1"] = "Requirement"
    ws["D1"] = "Response"
    ws["A2"] = "CTL-1"
    ws["B2"] = "Do you encrypt data at rest?"
    ws["D2"] = ""
    path = tmp_path / "custom.xlsx"
    wb.save(path)

    workbook = openpyxl.load_workbook(path)
    column_map = detect_columns(workbook.active, column_override={"question": 2, "answer": 4})
    assert column_map.question_col == 2
    assert column_map.answer_col == 4

    from src.questionnaire.parse_xlsx import _parse_column_override

    parsed = _parse_column_override("question=B,answer=D")
    assert parsed == {"question": 2, "answer": 4}
    assert _parse_column_override("question=2,answer=4,vocab=5")["vocab"] == 5
    import pytest

    with pytest.raises(ValueError):
        _parse_column_override("qustion=B,answer=D")
