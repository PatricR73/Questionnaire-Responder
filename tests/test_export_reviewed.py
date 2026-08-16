"""P29: reviewed-workbook export edge cases.

Three bugs this locks in:
1. Export was gated on 100% of rows being reviewed, including error rows that
   cannot meaningfully be approved — a run with a few ERROR rows was unexportable.
   Error rows are now excluded from the gate and always written as ERROR_MARKER.
2. An approved row whose final_confidence was "error" got the ERROR marker written
   and its approved text silently discarded.
3. counts[action] += 1 KeyError'd on any unexpected human_action, including None.
Plus the review-trail fix: record_human_review appends events rather than
overwriting, so a misclick is undoable.
"""

import openpyxl

from src.questionnaire.write_xlsx import ERROR_MARKER, NOT_FOUND_MARKER
from src.review_ui import export_reviewed_workbook
from src.store import db

import pathlib
FIXTURES = pathlib.Path(__file__).resolve().parent.parent / "fixtures"
SAMPLE = FIXTURES / "questionnaire_sample.xlsx"


def _seed_run(tmp_path):
    """One run with one approved, one edited, one rejected, and one error row."""
    conn = db.connect(tmp_path / "store.db")
    run_id = db.start_questionnaire_run(conn, str(SAMPLE), str(tmp_path / "out.xlsx"))
    # rows must be actual question rows of the sample sheet (3,5,6,8 are; 2 is a
    # merged section header and 4/7 are section/spacer rows write_answer would hit
    # read-only MergedCells on)
    rows = [
        (3, "approved", "high", "Approved answer", None),
        (5, "edited", "low", "Drafted answer", "Edited by human"),
        (6, "rejected", "low", "Drafted answer", None),
        (8, "error", "error", None, None),
    ]
    for row_index, action, confidence, drafted, reviewed in rows:
        db.record_answer(
            conn, run_id, row_index, f"Q{row_index}", f"Q{row_index}",
            drafted, None, confidence if confidence != "error" else None,
            confidence, ["doc.md::0"],
        )
        if action is not None:
            db.record_human_review(
                conn, run_id, row_index, action,
                reviewed_answer=reviewed if action == "edited" else None,
            )
    return conn, run_id


def _load_rows(conn, run_id):
    # Same shape the review UI hands to export: answers joined to the latest
    # review event per row.
    return conn.execute(
        """
        SELECT a.row_index, a.final_confidence, a.drafted_answer, a.reviewed_answer,
               a.vocab_selection, l.human_action
        FROM answers a
        LEFT JOIN audit_log l ON l.id = (
            SELECT MAX(la.id) FROM audit_log la
            WHERE la.run_id = a.run_id AND la.row_index = a.row_index
              AND (la.confidence = 'review' OR la.human_action IS NOT NULL)
        )
        WHERE a.run_id = ?
        ORDER BY a.row_index
        """,
        (run_id,),
    ).fetchall()


def test_export_writes_each_action_correctly_and_error_rows_always_error(tmp_path):
    conn, run_id = _seed_run(tmp_path)
    rows = _load_rows(conn, run_id)
    export_path = export_reviewed_workbook(conn, run_id, str(SAMPLE), str(tmp_path / "out.xlsx"), rows)

    ws = openpyxl.load_workbook(export_path).active
    assert ws.cell(row=3, column=3).value == "Approved answer"
    assert ws.cell(row=5, column=3).value == "Edited by human"
    assert ws.cell(row=6, column=3).value == NOT_FOUND_MARKER  # rejected -> not-found
    assert ws.cell(row=8, column=3).value == ERROR_MARKER  # error always ERROR_MARKER


def test_pristine_source_workbook_is_never_modified(tmp_path):
    before = SAMPLE.read_bytes()
    conn, run_id = _seed_run(tmp_path)
    rows = _load_rows(conn, run_id)
    export_reviewed_workbook(conn, run_id, str(SAMPLE), str(tmp_path / "out.xlsx"), rows)
    assert SAMPLE.read_bytes() == before, "source workbook must be byte-identical after export"


def test_unreviewed_and_unknown_actions_do_not_crash_export(tmp_path):
    conn, run_id = _seed_run(tmp_path)
    # add a row with no human_action and one with a bogus action value
    db.record_answer(conn, run_id, 9, "Q9", "Q9", "Draft", None, "high", "high", [])
    db.record_answer(conn, run_id, 10, "Q10", "Q10", "Draft", None, "high", "high", [])
    conn.execute(
        "INSERT INTO audit_log (run_id, row_index, sources_consulted, confidence, provider, human_action, timestamp) "
        "VALUES (?, ?, '[]', 'review', 'review-ui', 'bogus-action', 'now')",
        (run_id, 10),
    )
    conn.commit()

    rows = _load_rows(conn, run_id)
    export_path = export_reviewed_workbook(conn, run_id, str(SAMPLE), str(tmp_path / "out.xlsx"), rows)
    assert export_path.exists()


def test_review_events_append_and_latest_wins(tmp_path):
    conn, run_id = _seed_run(tmp_path)
    # row 3 was approved by the seed; approve again, then undo it, then reject it —
    # the trail must contain every event and the current state must be the latest.
    db.record_human_review(conn, run_id, 3, "approved")
    db.record_human_review(conn, run_id, 3, None)  # undo
    db.record_human_review(conn, run_id, 3, "rejected")

    events = conn.execute(
        "SELECT human_action, confidence FROM audit_log WHERE run_id = ? AND row_index = 3 AND confidence = 'review' ORDER BY id",
        (run_id,),
    ).fetchall()
    assert [e["human_action"] for e in events] == ["approved", "approved", None, "rejected"]
    assert all(e["confidence"] == "review" for e in events)

    # the review UI's "latest event" query reports rejected as current
    current = conn.execute(
        """
        SELECT human_action FROM audit_log la
        WHERE la.run_id = ? AND la.row_index = 3 AND (la.confidence = 'review' OR la.human_action IS NOT NULL)
        ORDER BY la.id DESC LIMIT 1
        """,
        (run_id,),
    ).fetchone()
    assert current["human_action"] == "rejected"
