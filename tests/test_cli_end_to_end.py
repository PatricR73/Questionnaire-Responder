"""P21: end-to-end CLI tests using the stub provider.

--provider stub and --stub-fail-row exist precisely so the whole pipeline can be
exercised for free, and nothing used them. These tests run the real CLI through
CliRunner — ingest, then answer — with no API key and no network beyond the
one-time embedding model download (already cached by test_pipeline_smoke).
"""

import os

import openpyxl
import pytest
from click.testing import CliRunner

from src.pipeline import STUB_BANNER_TEXT, cli
from src.questionnaire.write_xlsx import ERROR_FILL, ERROR_MARKER, NOT_FOUND_MARKER
from src.store import db

import pathlib
FIXTURES = pathlib.Path(__file__).resolve().parent.parent / "fixtures"
SAMPLE = FIXTURES / "questionnaire_sample.xlsx"
EVIDENCE = FIXTURES / "evidence"


def _ingest_and_answer(tmp_path, monkeypatch, answer_args):
    monkeypatch.setenv("QRESP_DATA_DIR", str(tmp_path / "data"))
    ingest = CliRunner().invoke(cli, ["ingest", "--evidence-dir", str(EVIDENCE)])
    assert ingest.exit_code == 0, ingest.output
    output = tmp_path / "out.xlsx"
    result = CliRunner().invoke(cli, ["answer", "--questionnaire", str(SAMPLE), "--output", str(output), *answer_args])
    return result, output


def test_stub_run_preserves_source_structure_and_populates_answers(tmp_path, monkeypatch):
    result, output = _ingest_and_answer(tmp_path, monkeypatch, ["--provider", "stub", "--limit", "0"])
    assert result.exit_code == 0, result.output

    source_wb = openpyxl.load_workbook(SAMPLE)
    out_wb = openpyxl.load_workbook(output)
    source_ws, out_ws = source_wb.active, out_wb.active

    # merged cells preserved — the source's merges must all survive; the only
    # addition allowed is the stub banner's own merge (A{last}:D{last})
    source_merges = {str(m) for m in source_ws.merged_cells.ranges}
    out_merges = {str(m) for m in out_ws.merged_cells.ranges}
    assert source_merges <= out_merges, f"source merges lost: {source_merges - out_merges}"
    assert out_merges - source_merges == {f"A{out_ws.max_row}:D{out_ws.max_row}"}

    # section-header and spacer rows untouched
    for row in (2, 4, 7):
        for col in range(1, source_ws.max_column + 1):
            assert source_ws.cell(row=row, column=col).value == out_ws.cell(row=row, column=col).value, f"row {row} col {col} changed"

    # every detected question row (3,5,6,8,9,10) has a populated answer cell
    for row in (3, 5, 6, 8, 9, 10):
        value = out_ws.cell(row=row, column=3).value
        assert value is not None and str(value).strip(), f"row {row} answer cell empty"
        assert value in (NOT_FOUND_MARKER, ERROR_MARKER) or value.startswith("[STUB]")


def test_stub_banner_is_merged_and_styled(tmp_path, monkeypatch):
    result, output = _ingest_and_answer(tmp_path, monkeypatch, ["--provider", "stub", "--limit", "0"])
    assert result.exit_code == 0, result.output

    ws = openpyxl.load_workbook(output).active
    banner_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == STUB_BANNER_TEXT:
            banner_row = row
            break
    assert banner_row is not None, "stub banner row must be present"

    merged = [m for m in ws.merged_cells.ranges if m.min_row == banner_row == m.max_row]
    assert merged, "banner must be merged across the used columns"
    assert merged[0].min_col == 1 and merged[0].max_col == 4, "banner spans question/answer/vocab columns"

    cell = ws.cell(row=banner_row, column=1)
    assert cell.fill.start_color.rgb == "00D32F2F"
    assert cell.font.bold is True
    assert cell.font.color.rgb == "00FFFFFF"


def test_stub_fail_row_marks_only_that_row(tmp_path, monkeypatch):
    result, output = _ingest_and_answer(tmp_path, monkeypatch, ["--provider", "stub", "--limit", "0", "--stub-fail-row", "5"])
    assert result.exit_code == 0, result.output

    ws = openpyxl.load_workbook(output).active
    assert ws.cell(row=5, column=3).value == ERROR_MARKER
    assert ws.cell(row=5, column=3).fill.start_color.rgb == ERROR_FILL.start_color.rgb
    for row in (3, 6, 8, 9, 10):
        assert ws.cell(row=row, column=3).value != ERROR_MARKER, f"row {row} must not be an error"

    # audit log: the error row and the others all recorded with provider=stub
    conn = db.connect(pathlib.Path(tmp_path) / "data" / "store.db")
    rows = conn.execute("SELECT row_index, confidence, provider FROM audit_log ORDER BY row_index").fetchall()
    assert len(rows) == 6
    assert all(r["provider"] == "stub" for r in rows)
    assert next(r for r in rows if r["row_index"] == 5)["confidence"] == "error"


def test_only_row_on_a_non_question_row_raises(tmp_path, monkeypatch):
    result, _ = _ingest_and_answer(tmp_path, monkeypatch, ["--provider", "stub", "--only-row", "4"])
    assert result.exit_code != 0
    assert "not a detected question row" in result.output


def test_limit_zero_processes_every_question_row(tmp_path, monkeypatch):
    result, _ = _ingest_and_answer(tmp_path, monkeypatch, ["--provider", "stub", "--limit", "0"])
    assert result.exit_code == 0, result.output
    conn = db.connect(pathlib.Path(tmp_path) / "data" / "store.db")
    assert conn.execute("SELECT COUNT(*) AS n FROM answers").fetchone()["n"] == 6


def test_anthropic_without_key_exits_nonzero_and_creates_no_run(tmp_path, monkeypatch):
    monkeypatch.setenv("QRESP_DATA_DIR", str(tmp_path / "data"))
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    output = tmp_path / "out.xlsx"
    result = CliRunner().invoke(
        cli,
        ["answer", "--questionnaire", str(SAMPLE), "--output", str(output), "--limit", "1", "--provider", "anthropic"],
    )
    assert result.exit_code != 0
    assert "ANTHROPIC_API_KEY" in result.output
    store = pathlib.Path(tmp_path) / "data" / "store.db"
    if store.exists():
        conn = db.connect(store)
        assert conn.execute("SELECT COUNT(*) AS n FROM questionnaire_runs").fetchone()["n"] == 0
