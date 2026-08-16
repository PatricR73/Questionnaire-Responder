"""P23: data-validation range matching.

The old membership test was `col_letter in str(rng)` — plain substring
containment — so column C matched a validation defined on AC1:AC10 and silently
picked up the wrong vocabulary list. Only inline list formulas were handled, not
range references (=Lists!$A$1:$A$3), and splitting on "," broke any value
containing a comma.
"""

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

from src.questionnaire.parse_xlsx import _vocab_values_for_column, detect_columns


def _sheet_with_validation(validation_range, formula):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Question"
    ws["B1"] = "Response"
    ws["C1"] = "Compliance Status"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(validation_range)
    ws.add_data_validation(dv)
    return ws


def test_column_c_does_not_match_a_validation_on_ac():
    # The regression: "C" is a substring of "AC1:AC10", so the old check made
    # column C adopt the AC column's vocabulary list.
    ws = _sheet_with_validation("AC1:AC10", '"Yes,No"')
    assert _vocab_values_for_column(ws, 3, 1) is None


def test_column_inside_the_range_gets_the_values():
    ws = _sheet_with_validation("C2:C100", '"Yes,No"')
    assert _vocab_values_for_column(ws, 3, 1) == ["Yes", "No"]


def test_range_reference_is_resolved_against_the_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Question"
    ws["B1"] = "Response"
    ws["C1"] = "Compliance Status"
    lists = wb.create_sheet("Lists")
    lists["A1"] = "Yes"
    lists["A2"] = "No"
    lists["A3"] = "N/A"
    dv = DataValidation(type="list", formula1="=Lists!$A$1:$A$3", allow_blank=True)
    dv.add("C2:C100")
    ws.add_data_validation(dv)
    assert _vocab_values_for_column(ws, 3, 1) == ["Yes", "No", "N/A"]


def test_quoted_values_with_commas_survive():
    # A value containing a comma is individually quoted; a naive split on ","
    # breaks it into two bogus values.
    ws = _sheet_with_validation("C2:C100", '"Yes, with exceptions","No"')
    assert _vocab_values_for_column(ws, 3, 1) == ["Yes, with exceptions", "No"]


def test_quoted_values_with_commas_survive_double_quoted_list():
    ws = _sheet_with_validation("C2:C100", '"Yes","No, not yet"')
    assert _vocab_values_for_column(ws, 3, 1) == ["Yes", "No, not yet"]


def test_detect_columns_uses_the_fixed_membership(tmp_path):
    # End to end: a vocab column whose dropdown is on a full-column range is
    # detected; the AC decoy is not.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Control ID"
    ws["B1"] = "Control Domain"
    ws["C1"] = "Question"
    ws["D1"] = "Response"
    ws["E1"] = "Compliance Status"
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv.add("E2:E100")
    ws.add_data_validation(dv)
    cm = detect_columns(ws)
    assert cm.question_col == 3
    assert cm.vocab_col == 5
    assert cm.vocab_values == ["Yes", "No"]
