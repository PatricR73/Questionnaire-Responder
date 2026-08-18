"""Tests for the gap report (pack 3, C5): NOT_FOUND rows as a deliverable.

A gap report is only useful if it distinguishes "nothing found" from "found
something adjacent but not on point", groups by domain, and reports the weak rows
too — without spending an API cent. These tests pin the structure on a small store.
"""

from pathlib import Path

import openpyxl
import pytest

from src.gap_report import build_gap_report, render_markdown, render_xlsx
from src.ingest.embed import ingest_evidence
from src.retrieval.hybrid_search import HybridSearcher
from src.store import db
from src.store.vectorstore import VectorStore

FIXTURES = Path(__file__).parent.parent / "fixtures"


@pytest.fixture()
def run_store(tmp_path):
    conn = db.connect(tmp_path / "store.db")
    vector_store = VectorStore(persist_dir=tmp_path / "chroma")
    ingest_evidence(FIXTURES / "evidence", conn, vector_store)
    run_id = db.start_questionnaire_run(conn, str(FIXTURES / "questionnaire_sample.xlsx"), "out/x.xlsx")
    # One NOT_FOUND row (no evidence), one low row.
    db.record_answer(
        conn, run_id, 2, "Do you offer a public bug bounty program?", "Do you offer a public bug bounty program?",
        None, None, None, "none", [],
    )
    db.record_audit_entry(conn, run_id, 2, [], "none", provider="stub")
    db.record_answer(
        conn, run_id, 3, "Is cloud data periodically backed up?", "Is cloud data periodically backed up?",
        "Yes, hourly, with a hedged caveat about retention windows.",
        None, "low", "low", [], polarity="partial",
    )
    db.record_audit_entry(conn, run_id, 3, ["business_continuity_plan.docx"], "low", provider="stub")
    return conn, run_id, vector_store


def test_gap_report_structure(run_store, tmp_path):
    conn, run_id, vector_store = run_store
    searcher = HybridSearcher(conn, vector_store)
    report = build_gap_report(conn, run_id, FIXTURES / "questionnaire_sample.xlsx", searcher)
    assert report["gap_count"] == 1
    assert report["weak_count"] == 1
    assert report["total_questions"] == 2
    gap = report["gaps"][0]
    assert gap.row_index == 2
    # Retrieval always returns its top-k, so a gap row reports WHAT was found and
    # how far it was — 'adjacent but not on point' — rather than claiming silence.
    assert gap.has_adjacent is True
    assert gap.closest_source is not None
    assert "distance" in gap.distance_label
    assert report["domains_ranked"]  # non-empty


def test_gap_report_marks_adjacent_retrieval(run_store, tmp_path):
    """A NOT_FOUND row whose retrieval DID find something must report the closest
    evidence and its distance — 'found something adjacent' is different from
    'nothing found' for the person fixing it."""
    conn, run_id, vector_store = run_store
    searcher = HybridSearcher(conn, vector_store)
    # An eval question that retrieves adjacent (but not on-point) evidence.
    run2 = db.start_questionnaire_run(conn, str(FIXTURES / "eval" / "questionnaire_eval.xlsx"), "out/y.xlsx")
    db.record_answer(
        conn, run2, 4, "Are reviews and revalidation of user access for least privilege completed?",
        "Are reviews and revalidation of user access for least privilege completed?",
        None, None, None, "none", [],
    )
    db.record_audit_entry(conn, run2, 4, [], "none", provider="stub")
    report = build_gap_report(conn, run2, FIXTURES / "eval" / "questionnaire_eval.xlsx", searcher)
    assert report["gap_count"] == 1
    gap = report["gaps"][0]
    assert gap.has_adjacent is True
    assert gap.closest_source is not None
    assert "distance" in gap.distance_label


def test_gap_report_renders_markdown_and_xlsx(run_store, tmp_path):
    conn, run_id, vector_store = run_store
    searcher = HybridSearcher(conn, vector_store)
    report = build_gap_report(conn, run_id, FIXTURES / "questionnaire_sample.xlsx", searcher)
    md = render_markdown(report)
    assert "Documentation gap report" in md
    assert "1 of 2 questions are unanswerable" in md
    assert "Documented but weakly supported" in md

    xlsx_path = tmp_path / "gap.xlsx"
    render_xlsx(report, xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    assert "Gaps" in wb.sheetnames
    assert "Weak (low confidence)" in wb.sheetnames
    assert wb["Gaps"].max_row >= 2  # header + at least one gap row
