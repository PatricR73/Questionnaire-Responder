"""write_answer output contract across all four final_confidence states.

Guards the P4 fix: the 'none' and 'error' paths must clear a pre-existing (customer
defaulted) vocabulary value — otherwise a sheet shipping with a defaulted compliance
column reads "Yes / NOT FOUND IN PROVIDED DOCUMENTS", a false representation to a
customer — and 'none' rows must get a visible neutral treatment (fill + comment) so
an honest abstention is not visually indistinguishable from a real answer in a long
sheet.
"""

import openpyxl

from src.questionnaire.parse_xlsx import ColumnMap
from src.questionnaire.write_xlsx import (
    ERROR_FILL,
    ERROR_MARKER,
    FLAG_FILL,
    NOT_FOUND_FILL,
    NOT_FOUND_MARKER,
    write_answer,
)


def _make_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Question"
    ws["B1"] = "Answer"
    ws["C1"] = "Compliance"
    column_map = ColumnMap(header_row=1, question_col=1, answer_col=2, vocab_col=3, vocab_values=["Yes", "No"])
    return ws, column_map


def _fill_sig(fill_obj):
    """Compare fills by colour + type, not by PatternFill object identity. Accepts
    either a Cell (read its .fill) or a PatternFill directly."""
    fill = fill_obj.fill if hasattr(fill_obj, "fill") else fill_obj
    return (fill.fill_type, fill.start_color.rgb if fill.start_color else None)


def test_high_confidence_row():
    ws, cm = _make_sheet()
    write_answer(ws, 2, cm, "Yes, we encrypt in transit.", "Yes", "high")

    assert ws["B2"].value == "Yes, we encrypt in transit."
    assert ws["B2"].comment is None
    # no special fill — a high-confidence answer is written plainly
    assert _fill_sig(ws["B2"]) not in (_fill_sig(FLAG_FILL), _fill_sig(NOT_FOUND_FILL), _fill_sig(ERROR_FILL))
    assert ws["C2"].value == "Yes"  # vocabulary value written through


def test_low_confidence_row():
    ws, cm = _make_sheet()
    write_answer(ws, 2, cm, "Partially — see the gap.", "No", "low")

    assert ws["B2"].value == "Partially — see the gap."
    assert _fill_sig(ws["B2"]) == _fill_sig(FLAG_FILL)
    assert ws["B2"].comment is not None
    assert "review" in ws["B2"].comment.text.lower()
    assert ws["C2"].value == "No"


def test_none_row_gets_neutral_fill_comment_and_cleared_vocab():
    ws, cm = _make_sheet()
    ws["C2"] = "Yes"  # pre-existing/defaulted customer value that must not survive
    write_answer(ws, 2, cm, "", None, "none")

    assert ws["B2"].value == NOT_FOUND_MARKER
    # neutral grey — not the low-confidence yellow, not the error red
    assert _fill_sig(ws["B2"]) == _fill_sig(NOT_FOUND_FILL)
    assert _fill_sig(ws["B2"]) != _fill_sig(FLAG_FILL)
    assert _fill_sig(ws["B2"]) != _fill_sig(ERROR_FILL)
    assert ws["B2"].comment is not None
    assert "human answer" in ws["B2"].comment.text
    assert ws["C2"].value is None  # stale default cleared


def test_error_row_gets_red_fill_comment_and_cleared_vocab():
    ws, cm = _make_sheet()
    ws["C2"] = "Yes"
    write_answer(ws, 2, cm, "", None, "error")

    assert ws["B2"].value == ERROR_MARKER
    assert _fill_sig(ws["B2"]) == _fill_sig(ERROR_FILL)
    assert ws["B2"].comment is not None
    assert ws["C2"].value is None  # stale default cleared


def test_none_and_error_clear_vocab_even_when_it_was_defaulted():
    for state in ("none", "error"):
        ws, cm = _make_sheet()
        ws["C2"] = "Applicable"
        write_answer(ws, 2, cm, "", None, state)
        assert ws["C2"].value is None, f"vocab must be cleared on {state}"


def test_high_low_leave_vocab_untouched_when_selection_is_none():
    # Out of scope for P4 (which clears only the none/error paths): when the model
    # returns no selection on a high/low row, the existing cell value is left alone.
    for state in ("high", "low"):
        ws, cm = _make_sheet()
        ws["C2"] = "Yes"
        write_answer(ws, 2, cm, "An answer.", None, state)
        assert ws["C2"].value == "Yes"
