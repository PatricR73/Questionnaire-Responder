"""Write drafted answers back into the original workbook, in place, touching only answer/vocab cells."""

from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.questionnaire.parse_xlsx import ColumnMap

NOT_FOUND_MARKER = "NOT FOUND IN PROVIDED DOCUMENTS"
ERROR_MARKER = "NOT PROCESSED — RUN FAILED, SEE AUDIT LOG"
FLAG_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # pale yellow
ERROR_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # pale red
# Neutral grey for an honest abstention — deliberately distinct from the
# low-confidence yellow (a real answer that needs review) and the error red (a row
# that was never processed). "Checked, no evidence found" is neither, and in a
# 300-row sheet it must be visually distinguishable from a real answer: a human
# reviewer has to find these rows on purpose, not by accident.
NOT_FOUND_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
# C4: third fill colour — a row whose answer drew on the answer library (a prior
# human-approved answer surfaced as a candidate). Pale green: distinct from the
# low-confidence yellow (a real answer needing review) and the NOT_FOUND grey, so a
# reviewer can see at a glance which rows came from the library versus fresh
# generation, and the comment carries the full provenance (source run, row, action,
# timestamp) so the reuse is auditable.
LIBRARY_FILL = PatternFill(start_color="D7E8D7", end_color="D7E8D7", fill_type="solid")
NOT_FOUND_COMMENT = Comment(
    "No supporting evidence was found in the provided documents for this question. "
    "This is an honest abstention, not a verified absence — the row needs a human answer.",
    "Questionnaire Responder",
)


def _clear_vocab_cell(ws: Worksheet, row_index: int, column_map: ColumnMap) -> None:
    """Blank any pre-existing vocabulary value on the none/error paths.

    If the customer's workbook ships with a pre-filled or defaulted compliance
    column and this row is left untouched, the output reads e.g. "Yes / NOT FOUND
    IN PROVIDED DOCUMENTS" — a false representation to a customer, emitted by the
    tool built specifically not to make one. A row with no answer (or a processing
    failure) must not carry any vocabulary value at all."""
    if column_map.vocab_col:
        ws.cell(row=row_index, column=column_map.vocab_col).value = None


def write_answer(
    ws: Worksheet,
    row_index: int,
    column_map: ColumnMap,
    answer_text: str,
    vocab_selection: str | None,
    final_confidence: str,
    library_state: str | None = None,
    library_provenance: str | None = None,
) -> None:
    """final_confidence is one of 'high', 'low', 'none' (checked, no evidence), or 'error'
    (not checked — a per-row failure, distinct from 'none' so it is never mistaken for a
    verified absence of evidence)."""
    answer_cell = ws.cell(row=row_index, column=column_map.answer_col)

    if final_confidence == "error":
        answer_cell.value = ERROR_MARKER
        answer_cell.fill = ERROR_FILL
        answer_cell.comment = Comment(
            "This row was not processed due to an error — re-run it, do not treat as 'no evidence'.",
            "Questionnaire Responder",
        )
        _clear_vocab_cell(ws, row_index, column_map)
        return

    if final_confidence == "none":
        answer_cell.value = NOT_FOUND_MARKER
        answer_cell.fill = NOT_FOUND_FILL
        answer_cell.comment = NOT_FOUND_COMMENT
        _clear_vocab_cell(ws, row_index, column_map)
        return

    answer_cell.value = answer_text
    if column_map.vocab_col and vocab_selection:
        ws.cell(row=row_index, column=column_map.vocab_col).value = vocab_selection

    if final_confidence == "low":
        answer_cell.fill = FLAG_FILL
        answer_cell.comment = Comment("Low confidence — needs human review before sending.", "Questionnaire Responder")

    # C4: the library marker. "used" rows get the green fill + provenance comment;
    # "surfaced" rows (a candidate existed but the generated answer did not
    # materially reuse it) get a comment only, so the marker never overrides the
    # confidence fill a reviewer actually needs to see.
    if library_state == "used":
        answer_cell.fill = LIBRARY_FILL
        answer_cell.comment = Comment(
            "Answer drew on a prior approved answer (answer library). " + (library_provenance or ""),
            "Questionnaire Responder",
        )
    elif library_state == "surfaced":
        answer_cell.comment = Comment(
            "A prior approved answer was surfaced for this question but the draft did not "
            "materially reuse it — verify against current evidence. " + (library_provenance or ""),
            "Questionnaire Responder",
        )
