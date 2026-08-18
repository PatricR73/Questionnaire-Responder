"""Detect question/answer/vocab columns in a vendor questionnaire workbook and read question rows.

Reads the workbook without altering it (write_xlsx.py owns writing back into the same
loaded Workbook object) so merged cells, section header rows, blank spacer rows, and
formatting stay untouched until an answer is actually written.
"""

import csv
import logging
from dataclasses import dataclass

from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger("qresp")

HEADER_SCAN_ROWS = 15

# "control" is in QUESTION_KEYWORDS because real sheets phrase the question column
# that way, but "Control ID" / "Control Domain" — which real CAIQ v4 sheets put
# BEFORE the question column — contain it too. First-match detection grabbed those
# (the original bug this file's tests pin). Scoring with decoy penalties fixes it.
QUESTION_KEYWORDS = ("question", "control", "requirement", "item", "criteria")
QUESTION_DECOYS = ("control id", "control domain")
ANSWER_KEYWORDS = ("answer", "response")
VOCAB_KEYWORDS = ("yes/no", "compliance", "applicable", "status", "y/n")


@dataclass
class ColumnMap:
    header_row: int
    question_col: int  # 1-based column index
    answer_col: int
    vocab_col: int | None
    vocab_values: list[str] | None


@dataclass
class QuestionRow:
    row_index: int  # 1-based row number in the sheet
    question_text: str


def _cell_text(ws: Worksheet, row: int, col: int) -> str:
    value = ws.cell(row=row, column=col).value
    return str(value).strip().lower() if value is not None else ""


def _score_column(text: str, keywords: tuple[str, ...], decoys: tuple[str, ...] = ()) -> int:
    """How strongly a header cell matches the keyword set.

    Exact matches are worth far more than substrings ('Question' is the question
    column; 'Questionnaire Metadata' merely mentions it), and known decoys
    subtract heavily — 'Control ID'/'Control Domain' both contain 'control' but
    are never the question column on a real CAIQ sheet."""
    score = 0
    for kw in keywords:
        if text == kw:
            score += 10
        elif kw in text:
            score += 1
    for decoy in decoys:
        if decoy in text:
            score -= 10
    return score


def _find_header_row(ws: Worksheet) -> int:
    for row in range(1, min(HEADER_SCAN_ROWS, ws.max_row) + 1):
        question_col = _find_column(ws, row, QUESTION_KEYWORDS, QUESTION_DECOYS)
        answer_col = _find_column(ws, row, ANSWER_KEYWORDS)
        if question_col is not None and answer_col is not None:
            return row
    raise ValueError(
        f"Could not find a header row with both a question and an answer column in the first {HEADER_SCAN_ROWS} rows"
    )


def _find_column(ws: Worksheet, header_row: int, keywords: tuple[str, ...], decoys: tuple[str, ...] = ()) -> int | None:
    """Best-scoring column in the header row for the keyword set, or None.

    First-match used to return the leftmost substring hit, which picked "Control
    ID" as the question column on real CAIQ sheets and one of several
    response-ish columns as the answer. The highest scorer wins; ties go to the
    leftmost column (stable iteration)."""
    best_col = None
    best_score = 0
    for col in range(1, ws.max_column + 1):
        text = _cell_text(ws, header_row, col)
        if not text:
            continue
        score = _score_column(text, keywords, decoys)
        if score > best_score:
            best_score = score
            best_col = col
    return best_col if best_score > 0 else None


def _parse_vocab_formula(formula: str, ws: Worksheet) -> list[str] | None:
    """Values from a list-validation formula, handling the two real formats:

    - inline lists: `"Yes,No"` (optionally with the list wrapped in double quotes,
      and values individually quoted so embedded commas survive — parsed with csv
      semantics, not a naive split on "," which breaks any value containing one);
    - range references: `=Lists!$A$1:$A$3` (common in enterprise templates),
      resolved against the workbook's sheets.
    """
    text = (formula or "").strip()
    if not text:
        return None
    if text.startswith("="):
        ref = text[1:]
        if "!" in ref:
            sheet_name, range_part = ref.split("!", 1)
            sheet_name = sheet_name.strip().strip("'")
            target = ws.parent[sheet_name] if sheet_name in ws.parent.sheetnames else None
            if target is None:
                return None
        else:
            target = ws
            range_part = ref
        values = []
        for row in target[range_part.replace("$", "")]:
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    values.append(str(cell.value).strip())
        return values or None
    if text.startswith('"') and text.endswith('"'):
        # CSV semantics first: individually-quoted values survive embedded commas
        # ("Yes","No, not yet" -> two values). If that yields a single field, the
        # whole list was one quoted string ("Yes,No") and the comma is the list
        # separator — split inside the quotes.
        fields = next(csv.reader([text]), [])
        if len(fields) > 1:
            return [f.strip() for f in fields if f.strip()] or None
        inner = text[1:-1]
        fields = next(csv.reader([inner]), [])
        return [f.strip() for f in fields if f.strip()] or None
    fields = next(csv.reader([text]), [])
    return [f.strip() for f in fields if f.strip()] or None


def _vocab_values_for_column(ws: Worksheet, col: int, header_row: int) -> list[str] | None:
    """Look for an openpyxl data validation (dropdown list) bound to this column.

    Membership is decided with the validation's actual cell-range math (rng.min_col
    / rng.max_col), not string containment: the old `col_letter in str(rng)` made
    column C match a validation defined on AC1:AC10, silently picking up the wrong
    vocabulary list. A validation bound to a full column (C:C) or any range that
    intersects the column counts."""
    target_col = column_index_from_string(ws.cell(row=header_row, column=col).column_letter)
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue
        if not any(rng.min_col <= target_col <= rng.max_col for rng in dv.sqref.ranges):
            continue
        values = _parse_vocab_formula(dv.formula1 or "", ws)
        if values:
            return values
    return None


def _score_columns(ws: Worksheet, header_row: int) -> dict:
    """Best-scoring (column, score) per role in the header row, for `qresp inspect`.

    Exposes the scoring that detect_columns already uses internally so a user can
    SEE what was chosen and why before spending a cent on a run (pack 3, C6) —
    detection is a heuristic, and the commercial failure mode is a prospect's first
    real file being mis-detected with no way to verify."""

    def best(keywords, decoys=()):
        best_col = None
        best_score = 0
        for col in range(1, ws.max_column + 1):
            text = _cell_text(ws, header_row, col)
            if not text:
                continue
            score = _score_column(text, keywords, decoys)
            if score > best_score:
                best_score = score
                best_col = col
        return (best_col, best_score) if best_col is not None else (None, 0)

    question_col, question_score = best(QUESTION_KEYWORDS, QUESTION_DECOYS)
    answer_col, answer_score = best(ANSWER_KEYWORDS)
    vocab_col, vocab_score = best(VOCAB_KEYWORDS)
    return {
        "question": (question_col, question_score),
        "answer": (answer_col, answer_score),
        "vocab": (vocab_col, vocab_score),
    }


def _parse_column_override(raw: str | None) -> dict | None:
    """Parse --map question=C,answer=E,vocab=D into {role: 1-based column int}.

    Values may be letters (C, AB) or numbers (3, 28). Unknown roles raise, so a
    typo fails loudly instead of silently leaving one role auto-detected."""
    if not raw:
        return None
    result = {}
    for part in raw.split(","):
        if "=" not in part:
            raise ValueError(
                f"--map entries must be role=column, got {part!r} (valid roles: question, answer, vocab)"
            )
        role, value = part.strip().split("=", 1)
        role = role.strip().lower()
        if role not in ("question", "answer", "vocab"):
            raise ValueError(f"Unknown --map role {role!r} (valid: question, answer, vocab)")
        value = value.strip()
        if value.isdigit():
            result[role] = int(value)
        else:
            from openpyxl.utils.cell import column_index_from_string

            result[role] = column_index_from_string(value.upper())
    return result


def detect_columns(ws: Worksheet, column_override: dict | None = None) -> ColumnMap:
    """Detect the question/answer/vocab columns, optionally overridden by --map
    (pack 3, C6): a mis-detected sheet is a dead end for a prospect, and the
    override turns that dead end into a one-flag fix. Override keys are the roles
    ('question', 'answer', 'vocab') mapped to 1-based column indices."""
    header_row = _find_header_row(ws)
    question_col = _find_column(ws, header_row, QUESTION_KEYWORDS, QUESTION_DECOYS)
    answer_col = _find_column(ws, header_row, ANSWER_KEYWORDS)
    if column_override:
        question_col = column_override.get("question", question_col)
        answer_col = column_override.get("answer", answer_col)
        if question_col is None or answer_col is None:
            raise ValueError("--map must provide at least question and answer columns")
    if question_col is None or answer_col is None:
        raise ValueError(f"Header row {header_row} is missing a question or answer column")

    vocab_col = _find_column(ws, header_row, VOCAB_KEYWORDS)
    if column_override and "vocab" in column_override:
        vocab_col = column_override["vocab"]
    vocab_values = _vocab_values_for_column(ws, vocab_col, header_row) if vocab_col else None

    log.info(
        "Detected columns: question=%s answer=%s vocab=%s (header row %s) — check before spending money on a run.",
        question_col,
        answer_col,
        vocab_col,
        header_row,
    )
    return ColumnMap(
        header_row=header_row,
        question_col=question_col,
        answer_col=answer_col,
        vocab_col=vocab_col,
        vocab_values=vocab_values,
    )


def iter_question_sheets(workbook, column_override: dict | None = None) -> list[tuple[str, Worksheet, ColumnMap]]:
    """Every sheet in the workbook that carries a detectable question/answer pair.

    Real CAIQ v4.0.2 workbooks are multi-tab; SIG Lite is multi-tab; customer
    sheets routinely put instructions on tab 1 and questions on tab 2. Processing
    only workbook.active silently ignored everything but the first tab — the most
    likely single point of first-contact failure for a new user (pack 3, C6).
    Sheets without a detectable header are skipped, not crashed on."""
    found = []
    for ws in workbook.worksheets:
        try:
            column_map = detect_columns(ws, column_override=column_override)
        except ValueError:
            log.info("sheet %r has no detectable question/answer header — skipped", ws.title)
            continue
        found.append((ws.title, ws, column_map))
    return found





def _is_section_header_row(ws: Worksheet, row: int, question_col: int) -> bool:
    """A row is a section header (not a real question) if its question cell is merged across multiple columns."""
    cell = ws.cell(row=row, column=question_col)
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range and (merged_range.max_col - merged_range.min_col) > 0:
            return True
    return False


def read_questions(ws: Worksheet, column_map: ColumnMap) -> list[QuestionRow]:
    questions: list[QuestionRow] = []
    for row in range(column_map.header_row + 1, ws.max_row + 1):
        value = ws.cell(row=row, column=column_map.question_col).value
        text = str(value).strip() if value is not None else ""
        if not text:
            continue  # blank spacer row
        if _is_section_header_row(ws, row, column_map.question_col):
            continue
        questions.append(QuestionRow(row_index=row, question_text=text))
    return questions
