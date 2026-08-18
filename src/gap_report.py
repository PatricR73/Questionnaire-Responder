"""Gap report: the NOT_FOUND and low-confidence rows, as a deliverable.

Pack 3, C5. Every NOT_FOUND row is individually a blank cell and a marker string;
collectively they are a documentation gap analysis — "your policy set does not
document: business continuity test cadence, sub-processor inventory, encryption key
rotation, incident notification SLA." That reframes the pitch from "fill this sheet
faster" (a productivity tweak) to "find out what you need to write before a
customer asks" (risk management).

The data is already in the store: the answers table holds every row's status, and
retrieval is deterministic and free, so for each NOT_FOUND row this report re-runs
the local searcher (no API calls, no model beyond the cached local embedding) to
distinguish "nothing found" from "found something adjacent but not on point" — very
different findings for the person who has to fix it.

Outputs: a Markdown report and an XLSX workbook (one row per gap, grouped by
questionnaire domain where the sheet provides one; low-confidence rows as a second
section).
"""

import sqlite3
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

NOT_FOUND = "none"
LOW = "low"

_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_GAP_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


@dataclass
class GapRow:
    row_index: int
    question_text: str
    domain: str
    closest_source: str | None
    closest_heading: str | None
    closest_distance: float | None
    has_adjacent: bool  # something was retrieved at all
    distance_label: str  # "none retrieved" | "adjacent (distance X)" | "on-point?"


def _domain_for(workbook_sheet, row_index: int, header_row: int) -> str:
    """Domain from the sheet's first column (Control ID on CAIQ-style sheets), or
    "Uncategorized" when the column doesn't look like an ID. An ID looks like
    "BCR-08.1" — a short token containing a hyphen. Long prose or empty cells fall
    back to Uncategorized."""
    value = workbook_sheet.cell(row=row_index, column=1).value
    text = str(value).strip() if value is not None else ""
    if text and "-" in text and len(text) <= 16:
        return text.split("-")[0]
    return "Uncategorized"


def _closest_retrieved(searcher, question_text: str, top_k: int) -> GapRow:
    """Re-run local retrieval (free, deterministic) to characterise the gap:
    what was the closest evidence actually retrieved, and how far away was it?"""
    try:
        hits = searcher.search(question_text, top_k=top_k)
    except Exception:  # noqa: BLE001 — retrieval failure must not kill the report
        hits = []
    if not hits:
        return GapRow(0, "", "", None, None, None, False, "none retrieved")
    best = min(hits, key=lambda c: c.vector_distance if c.vector_distance is not None else float("inf"))
    distance = best.vector_distance
    label = f"adjacent (distance {distance:.3f})" if distance is not None else "retrieved (no distance)"
    return GapRow(0, "", "", best.source_filename, best.heading_path, distance, True, label)


def build_gap_report(
    conn: sqlite3.Connection,
    run_id: int,
    questionnaire_path: Path,
    searcher,
    *,
    top_k: int = 5,
) -> dict:
    """Gather the report data: gaps (NOT_FOUND) and weak rows (low), each with the
    closest evidence actually retrieved and its distance, grouped by domain."""
    source_row = conn.execute("SELECT source_path FROM questionnaire_runs WHERE id = ?", (run_id,)).fetchone()
    if source_row is None:
        raise ValueError(f"No questionnaire run with id {run_id}")

    rows = conn.execute(
        "SELECT row_index, question_text, final_confidence, polarity, drafted_answer "
        "FROM answers WHERE run_id = ? ORDER BY row_index",
        (run_id,),
    ).fetchall()

    # Domain lookup from the source workbook's first column (Control IDs).
    ws = None
    try:
        wb = openpyxl.load_workbook(questionnaire_path)
        ws = wb.active
    except Exception:  # noqa: BLE001 — workbook unreadable: everything is Uncategorized
        ws = None

    gaps: list[GapRow] = []
    weak: list[dict] = []
    for r in rows:
        row_index = r["row_index"]
        domain = _domain_for(ws, row_index, header_row=1) if ws is not None else "Uncategorized"
        if r["final_confidence"] == NOT_FOUND:
            closest = _closest_retrieved(searcher, r["question_text"], top_k)
            gaps.append(
                GapRow(
                    row_index,
                    r["question_text"],
                    domain,
                    closest.closest_source,
                    closest.closest_heading,
                    closest.closest_distance,
                    closest.has_adjacent,
                    closest.distance_label,
                )
            )
        elif r["final_confidence"] == LOW:
            weak.append(
                {
                    "row_index": row_index,
                    "question_text": r["question_text"],
                    "domain": domain,
                    "polarity": r["polarity"],
                    "drafted_answer": r["drafted_answer"] or "",
                }
            )

    domains: dict[str, int] = {}
    for g in gaps:
        domains[g.domain] = domains.get(g.domain, 0) + 1
    ranked = sorted(domains.items(), key=lambda kv: (-kv[1], kv[0]))
    return {
        "run_id": run_id,
        "source_path": source_row["source_path"],
        "total_questions": len(rows),
        "gaps": gaps,
        "weak": weak,
        "gap_count": len(gaps),
        "weak_count": len(weak),
        "domains_ranked": ranked,
    }


def render_markdown(report: dict) -> str:
    g = report["gaps"]
    lines = []
    lines.append(f"# Documentation gap report — run {report['run_id']}")
    lines.append("")
    lines.append(f"Source questionnaire: `{report['source_path']}`")
    lines.append("")
    lines.append(
        f"**{report['gap_count']} of {report['total_questions']} questions are unanswerable from the "
        f"current documentation** (plus {report['weak_count']} answered but flagged low-confidence). "
        "Each gap below shows the closest evidence retrieval actually found and its cosine distance, "
        "so 'nothing found' and 'found something adjacent but not on point' are distinguishable."
    )
    lines.append("")
    lines.append("## Most affected domains")
    lines.append("")
    lines.append("| Domain | Gaps |")
    lines.append("|---|---|")
    for domain, count in report["domains_ranked"]:
        lines.append(f"| {domain} | {count} |")
    lines.append("")

    lines.append("## Unanswerable questions (NOT FOUND)")
    lines.append("")
    lines.append("| Row | Domain | Question | Closest evidence retrieved | Distance |")
    lines.append("|---|---|---|---|---|")
    for g_row in g:
        src = g_row.closest_source or "—"
        head = f" / {g_row.closest_heading}" if g_row.closest_heading else ""
        lines.append(
            f"| {g_row.row_index} | {g_row.domain} | {g_row.question_text.replace(chr(10), ' ')[:100]} "
            f"| {src}{head} | {g_row.distance_label} |"
        )
    lines.append("")

    if report["weak"]:
        lines.append("## Documented but weakly supported (low confidence)")
        lines.append("")
        lines.append("| Row | Domain | Question | Polarity |")
        lines.append("|---|---|---|---|")
        for w_row in report["weak"]:
            lines.append(
                f"| {w_row['row_index']} | {w_row['domain']} | {w_row['question_text'].replace(chr(10), ' ')[:100]} | {w_row['polarity'] or '—'} |"
            )
        lines.append("")
    lines.append(
        "_Generated by `qresp gap-report` — retrieval-only, no API calls, no model beyond the local embedding._"
    )
    return chr(10).join(lines)


def render_xlsx(report: dict, out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gaps"
    ws.append(["Row", "Domain", "Question", "Closest evidence", "Heading", "Distance", "Adjacent found"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    for g in report["gaps"]:
        ws.append(
            [
                g.row_index,
                g.domain,
                g.question_text,
                g.closest_source or "",
                g.closest_heading or "",
                g.distance_label,
                "yes" if g.has_adjacent else "no",
            ]
        )
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            if cell.row % 2 == 0:
                cell.fill = _GAP_FILL
    for i, width in enumerate([6, 14, 60, 30, 30, 22, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws2 = wb.create_sheet("Weak (low confidence)")
    ws2.append(["Row", "Domain", "Question", "Polarity", "Drafted answer"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    for w_row in report["weak"]:
        ws2.append(
            [
                w_row["row_index"],
                w_row["domain"],
                w_row["question_text"],
                w_row["polarity"] or "",
                w_row["drafted_answer"],
            ]
        )
    for i, width in enumerate([6, 14, 60, 12, 80], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = width
    ws2.freeze_panes = "A2"
    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
