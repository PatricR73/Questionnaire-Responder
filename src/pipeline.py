"""CLI entrypoint: `ingest` an evidence directory, then `answer` a questionnaire against it.

Usage:
    python -m src.pipeline ingest --evidence-dir fixtures/evidence/
    python -m src.pipeline answer --questionnaire fixtures/questionnaire_sample.xlsx --output out/filled.xlsx --limit 5
    python -m src.pipeline answer --questionnaire fixtures/questionnaire_sample.xlsx --output out/filled.xlsx --provider stub

`answer` always saves the workbook on exit — including on an unhandled exception — via
a try/finally, and additionally saves every SAVE_EVERY_N_ROWS rows so progress is
visible on disk during a long run. The finally save is what guarantees correctness
(the xlsx always reflects everything processed so far, matching the sidecar .jsonl
log); the periodic save is purely a "don't wait until the end to see progress"
convenience on top of that guarantee, not a substitute for it. (The one thing neither
can help with is a hard kill, e.g. SIGKILL — no process can run cleanup code after
that; only a caught exception or a normal/Ctrl-C exit triggers the finally save.)

Answering is delegated to an Answerer (src/answer/answerer.py) selected via --provider:
"anthropic" (default, real Claude calls) or "stub" (no network/model — for exercising
the pipeline's plumbing and failure paths for free). A missing API key with
--provider anthropic always errors; there is no automatic fallback to stub. Stub runs
are stamped into the audit log and get a visible banner row in the output workbook so
a stub file can never be mistaken for a real one.

A per-row exception (from either Answerer) is caught here — not inside the Answerer —
and recorded as AnswerStatus.ERROR, written to the cell as a distinct state (never
conflated with a verified "no evidence found"; see write_xlsx.ERROR_MARKER), and the
loop continues to the next row.
"""

import json
import logging
import os
import re
import time
from datetime import UTC, datetime
from pathlib import Path

import anthropic
import click
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from src.answer.answerer import Answerer, AnswerStatus, AnthropicAnswerer, StubAnswerer
from src.answer.split_questions import split_question
from src.ingest.embed import ingest_evidence
from src.questionnaire.parse_xlsx import detect_columns, read_questions
from src.questionnaire.write_xlsx import write_answer
from src.retrieval.hybrid_search import HybridSearcher, RetrievedChunk
from src.store import db
from src.store.vectorstore import VectorStore

SAVE_EVERY_N_ROWS = 5

log = logging.getLogger("qresp")


class _JsonLinesFormatter(logging.Formatter):
    """One JSON object per line, carrying the structured per-row data attached as
    the 'row_data' LogRecord attribute plus the message, level, and traceback for
    failed rows — the file a debugger wants to read instead of re-running a 400-row
    sheet with --only-row and hoping a bad row reproduces."""

    def format(self, record: logging.LogRecord) -> str:
        payload = {
            "ts": datetime.now(UTC).isoformat(),
            "level": record.levelname,
            "logger": record.name,
            "message": record.getMessage(),
        }
        row_data = getattr(record, "row_data", None)
        if row_data:
            payload.update(row_data)
        if record.exc_info:
            payload["traceback"] = self.formatException(record.exc_info)
        return json.dumps(payload)


def _setup_logging(output: Path, verbose: bool, quiet: bool) -> None:
    """Configure the qresp logger: human-readable progress on stderr (INFO; DEBUG
    with --verbose; only errors with --quiet), and a JSON-lines file next to the
    output workbook at DEBUG with every structured per-row record. The existing
    click.echo progress lines are intentionally kept — they are the interactive UX;
    the logger's stderr channel is the --verbose/--quiet-controlled complement, and
    the file is the structured record."""
    log.setLevel(logging.DEBUG)
    log.handlers.clear()
    log.propagate = False

    stderr = logging.StreamHandler()
    stderr.setLevel(logging.DEBUG if verbose else (logging.ERROR if quiet else logging.INFO))
    stderr.setFormatter(logging.Formatter("%(levelname)s %(message)s"))
    log.addHandler(stderr)

    jsonl_path = output.with_suffix(".log.jsonl")
    file_handler = logging.FileHandler(jsonl_path, mode="a")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(_JsonLinesFormatter())
    log.addHandler(file_handler)


# Errors that will fail EVERY row identically and are never worth a per-row retry
# ladder: a 401 (AuthenticationError), a permission problem, a wrong model name
# (NotFoundError), or a schema-rejecting 400 (BadRequestError). Caught per-row
# today, each of these burns the full retry budget on every one of 400 rows and
# produces a workbook of red ERROR cells instead of stopping in the first ten
# seconds. They propagate out of the per-row handler and abort the run — the
# finally save still writes everything already processed, so no paid-for work is
# lost. Anything not in this set is treated as transient per-row noise.
FATAL_ERRORS = (
    anthropic.AuthenticationError,
    anthropic.PermissionDeniedError,
    anthropic.NotFoundError,
    anthropic.BadRequestError,
)

# Circuit breaker: N consecutive per-row errors (of any kind) abort the run. A
# systemic failure that isn't in FATAL_ERRORS — e.g. a network path that breaks for
# everyone, or an environment issue — would otherwise burn the whole sheet one row
# at a time; a healthy run never sees anywhere near this many in a row.
CONSECUTIVE_ERROR_LIMIT = 5

STUB_BANNER_TEXT = "⚠ STUB PROVIDER — THESE ARE NOT REAL ANSWERS (TESTING ONLY)"
STUB_BANNER_FILL = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
STUB_BANNER_FONT = Font(bold=True, color="FFFFFF")


@click.group()
@click.option(
    "--workspace",
    type=str,
    default=None,
    help="Namespaced data directory for this command (pack 3, C9): each workspace gets "
    "its own SQLite store, Chroma index, and answer library under "
    "<data_dir>/workspaces/<name>/, so a consultant/vCISO/MSP can keep clients "
    "structurally separate — client A's evidence and approved answers can never be "
    "retrieved for client B. Isolation is enforced at the storage layer (a different "
    "data directory per workspace), not by filtering: retrieval code never sees "
    "another workspace's rows. See 'qresp workspace list' / 'qresp workspace new'.",
)
@click.pass_context
def cli(ctx, workspace):
    ctx.ensure_object(dict)
    ctx.obj["workspace"] = workspace
    if workspace:
        from src.data_dir import REPO_ROOT

        base = Path(os.environ.get("QRESP_DATA_DIR") or REPO_ROOT / "out")
        ws_dir = base / "workspaces" / workspace
        ws_dir.mkdir(parents=True, exist_ok=True)
        # Set BEFORE any command body runs: db.connect() and VectorStore resolve
        # data_dir() at call time, so the env var alone is the isolation mechanism.
        os.environ["QRESP_DATA_DIR"] = str(ws_dir)
        log.info("workspace %r -> %s", workspace, ws_dir)


@cli.command()
@click.option("--evidence-dir", type=click.Path(exists=True, file_okay=False, path_type=Path), required=True)
def ingest(evidence_dir: Path):
    click.echo(f"Ingesting evidence from {evidence_dir} ...")
    click.echo("(first run downloads the local embedding model — this can take a few minutes and looks like a hang)")
    conn = db.connect()
    vector_store = VectorStore()
    n = ingest_evidence(evidence_dir, conn, vector_store)
    click.echo(f"Ingested {n} chunks.")



@cli.command()
@click.option(
    "--questionnaire",
    type=click.Path(exists=True, dir_okay=False, path_type=Path),
    default=None,
    help="Questionnaire to demo (default: the committed 24-question eval workbook).",
)
@click.option(
    "--port",
    type=int,
    default=8501,
    show_default=True,
    help="Port for the review screen launched at the end.",
)
@click.option(
    "--no-ui",
    is_flag=True,
    help="Fill the demo workbook and print paths, but do not launch the review screen.",
)
@click.option(
    "--bind",
    type=str,
    default="127.0.0.1",
    show_default=True,
    help="Address the review screen binds to. Keep 127.0.0.1 on a bare machine (the "
    "UI serves a database of internal policy text); the Docker image uses 0.0.0.0 "
    "because the container's network namespace is already the boundary there and "
    "the port must be reachable through the -p mapping.",
)
def demo(questionnaire: Path | None, port: int, no_ui: bool, bind: str):
    """One command, no API key, no ingest: a filled workbook plus a running review screen.

    Commercial rationale (pack 3, C1): the strongest evidence this project has — the
    measured eval, the citation audit trail, the honest NOT_FOUND behaviour — used to
    be gated behind a twenty-minute setup (clone, venv, ~1 GB of dependencies, a
    few-hundred-MB model download, an API key, two CLI commands). Almost nobody
    completed it. This command collapses that to one step against a PRE-BUILT store
    (committed under demo_store/ and built from the synthetic fixtures/evidence/),
    so ingest — and therefore the embedding model — is not required, and
    --provider stub means no Anthropic key and no API spend. The one remaining
    one-time download is the local embedding model needed to embed the QUERY side of
    retrieval; the Docker image (see the README) bakes that model in at build time,
    so a bare "docker run" is truly zero-download.

    The demo store is copied to out/demo_store/ before the run so the committed
    snapshot stays pristine; the run's workbook and sidecar land in out/demo/.
    The review screen is launched over that copy, bound to localhost.
    """
    import shutil
    import subprocess
    import sys

    from src.data_dir import REPO_ROOT

    store_src = REPO_ROOT / "demo_store"
    if not (store_src / "store.db").exists():
        raise click.ClickException(
            f"Demo store not found at {store_src}. This build ships without the pre-built "
            "demo store; run 'qresp ingest --evidence-dir fixtures/evidence' and re-run this command."
        )
    questionnaire_path = questionnaire or (REPO_ROOT / "fixtures" / "eval" / "questionnaire_eval.xlsx")
    if not questionnaire_path.exists():
        raise click.ClickException(f"Demo questionnaire not found at {questionnaire_path}.")

    demo_store_dir = REPO_ROOT / "out" / "demo_store"
    shutil.rmtree(demo_store_dir, ignore_errors=True)
    shutil.copytree(store_src, demo_store_dir)

    out_dir = REPO_ROOT / "out" / "demo"
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / "demo_filled.xlsx"

    click.echo("Running the demo questionnaire against the pre-built store with --provider stub ...")
    click.echo("(first run may download the local embedding model — a few hundred MB, one time;")
    click.echo(" the Docker image bakes it in, so a bare 'docker run' skips this entirely)")
    env = dict(os.environ)
    env["QRESP_DATA_DIR"] = str(demo_store_dir)
    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "src.pipeline",
            "answer",
            "--questionnaire",
            str(questionnaire_path),
            "--output",
            str(output_path),
            "--limit",
            "0",
            "--provider",
            "stub",
        ],
        env=env,
    )
    if result.returncode != 0:
        raise click.ClickException(f"Demo run failed (exit {result.returncode}).")
    click.echo("")
    click.echo(f"Demo workbook:  {output_path}")
    click.echo(f"Run sidecar:    {output_path.with_suffix('.jsonl')}")
    click.echo(f"Demo store:     {demo_store_dir}")
    if no_ui:
        click.echo("Review screen skipped (--no-ui).")
        return

    import src.review_ui as review_ui_module

    # The review UI reads its data dir at call time, so the env var set here is
    # picked up by the in-process launch below.
    os.environ["QRESP_DATA_DIR"] = str(demo_store_dir)
    click.echo(f"Launching the review screen at http://localhost:{port} (Ctrl-C to stop) ...")
    from streamlit.web import cli as stcli

    sys.argv = [
        "streamlit",
        "run",
        str(Path(review_ui_module.__file__).resolve()),
        "--server.address",
        bind,
        "--server.port",
        str(port),
        "--server.headless",
        "true",
    ]
    stcli.main()








@cli.command()
@click.argument("action", type=click.Choice(["list", "new"]))
@click.argument("name", type=str, required=False)
def workspace(action: str, name: str | None):
    """List workspaces, or create a new one (pack 3, C9).

    Workspaces are namespaced data directories — each has its own SQLite store,
    Chroma index, and answer library, so cross-client contamination is structurally
    impossible rather than merely unlikely. The CLI's --workspace option runs any
    command inside a workspace; this command manages them.

    qresp workspace list          # show every workspace and its data directory
    qresp workspace new acme      # create the acme workspace
    """
    from src.data_dir import REPO_ROOT

    base = Path(os.environ.get("QRESP_DATA_DIR") or REPO_ROOT / "out") / "workspaces"
    if action == "list":
        if not base.exists():
            click.echo("No workspaces yet — create one with 'qresp workspace new <name>'.")
            return
        for child in sorted(base.iterdir()):
            if child.is_dir():
                click.echo(f"{child.name}  ->  {child}")
        return
    if not name:
        raise click.ClickException("qresp workspace new needs a workspace name.")
    if not re.fullmatch(r"[a-zA-Z0-9_-]+", name):
        raise click.ClickException("Workspace names may contain only letters, digits, - and _.")
    target = base / name
    target.mkdir(parents=True, exist_ok=True)
    click.echo(f"Created workspace {name!r} at {target}")
    click.echo("Use it with: qresp --workspace <name> ingest --evidence-dir ... && qresp --workspace <name> answer ...")


@cli.command()
@click.option("--run-id", type=int, required=True, help="Questionnaire run id (see the review UI sidebar or store.db).")
@click.option(
    "--output-dir",
    type=click.Path(path_type=Path),
    default=None,
    help="Directory for the .md and .xlsx reports (default: out/).",
)
@click.option("--top-k", type=int, default=5, show_default=True, help="Retrieved chunks inspected per gap row.")
def gap_report(run_id: int, output_dir: Path | None, top_k: int):
    """Turn a run's NOT_FOUND rows into a documentation gap analysis (pack 3, C5).

    Every NOT_FOUND row is individually a blank cell; collectively they are a
    deliverable — "your policy set does not document: X, Y, Z." For each gap this
    command re-runs LOCAL retrieval (deterministic, zero API calls, no model beyond
    the cached local embedding) to distinguish 'nothing found' from 'found something
    adjacent but not on point', groups gaps by questionnaire domain, and writes a
    Markdown report plus an XLSX workbook. Low-confidence rows are reported as a
    second section: documented but weakly supported.
    """
    from src.config import load_config
    from src.data_dir import REPO_ROOT
    from src.gap_report import build_gap_report, render_markdown, render_xlsx

    conn = db.connect()
    src_row = conn.execute(
        "SELECT source_path FROM questionnaire_runs WHERE id = ?", (run_id,)
    ).fetchone()
    if src_row is None:
        raise click.ClickException(f"No questionnaire run with id {run_id}.")

    questionnaire_path = Path(src_row["source_path"])
    if not questionnaire_path.is_absolute():
        candidate = REPO_ROOT / questionnaire_path
        if candidate.exists():
            questionnaire_path = candidate
    if not questionnaire_path.exists():
        click.echo(f"(questionnaire workbook not found at {questionnaire_path} — domains will be Uncategorized)")

    cfg = load_config()
    vector_store = VectorStore(model_name=cfg.embedding_model)
    searcher = HybridSearcher(
        conn,
        vector_store,
        vector_weight=cfg.vector_weight,
        rrf_k=cfg.rrf_k,
        candidate_pool=cfg.candidate_pool,
    )
    report = build_gap_report(conn, run_id, questionnaire_path if questionnaire_path.exists() else Path(""), searcher, top_k=top_k)

    out_dir = output_dir or (REPO_ROOT / "out")
    out_dir.mkdir(parents=True, exist_ok=True)
    md_path = out_dir / f"gap_report_run{run_id}.md"
    xlsx_path = out_dir / f"gap_report_run{run_id}.xlsx"
    md_path.write_text(render_markdown(report))
    render_xlsx(report, xlsx_path)
    click.echo(f"Wrote {md_path}")
    click.echo(f"Wrote {xlsx_path}")
    click.echo(
        f"{report['gap_count']} unanswerable + {report['weak_count']} low-confidence of "
        f"{report['total_questions']} questions. No API calls were made."
    )



@cli.command()
@click.option("--questionnaire", type=click.Path(exists=True, dir_okay=False, path_type=Path), required=True)
@click.option(
    "--sheet",
    type=str,
    default=None,
    help="Inspect only this worksheet by name (default: every sheet).",
)
def inspect(questionnaire: Path, sheet: str | None):
    """Show what the pipeline will detect in a questionnaire — free, no API calls.

    Pack 3, C6: column detection is a heuristic, and the commercial failure mode is
    a prospect's first real file (multi-tab CAIQ, an instructions tab first, a
    bespoke layout) being mis-detected with no way to verify before spending money.
    This prints, per sheet: the detected header row, the chosen question/answer/
    vocab columns WITH their detection scores, the question-row count, and a sample
    question. Sheets without a detectable header are listed as skipped, so a silent
    multi-tab miss becomes visible in one command. Combine with --map on the answer
    command when detection guesses wrong.
    """
    import openpyxl

    from src.questionnaire.parse_xlsx import _score_columns, detect_columns, read_questions

    workbook = openpyxl.load_workbook(questionnaire, read_only=False)
    click.echo(f"Inspecting {questionnaire} ({len(workbook.sheetnames)} sheet(s)) — no API calls, nothing written.")
    for ws in workbook.worksheets:
        if sheet is not None and ws.title != sheet:
            continue
        click.echo("")
        click.echo(f"=== Sheet: {ws.title!r} (active={ws.title == workbook.active.title}) ===")
        try:
            column_map = detect_columns(ws)
            scores = _score_columns(ws, column_map.header_row)
        except ValueError as exc:
            click.echo(f"  SKIPPED — {exc}")
            continue
        click.echo(f"  header row: {column_map.header_row}")
        for role, (col, score) in scores.items():
            if col is None:
                click.echo(f"  {role:8s}: none (best score {score})")
            else:
                from openpyxl.utils.cell import get_column_letter

                header_text = ws.cell(row=column_map.header_row, column=col).value
                click.echo(
                    f"  {role:8s}: column {get_column_letter(col)} ({col}) score={score} "
                    f"header={header_text!r}"
                )
        questions = read_questions(ws, column_map)
        click.echo(f"  question rows: {len(questions)}")
        if questions:
            click.echo(f"  sample: {questions[0].question_text[:110]}")
    click.echo("")
    click.echo("If detection is wrong on a real file, re-run answer with --map question=C,answer=E,vocab=D.")



@cli.command()
@click.option("--port", type=int, default=8000, show_default=True)
@click.option("--host", type=str, default="127.0.0.1", show_default=True)
def serve(port: int, host: str):
    """Run the integration HTTP service (pack 3, C10).

    Submit a questionnaire run, poll its status, fetch its structured results —
    the three operations an intake flow needs. Thin HTTP wrapper over the
    qresp.Pipeline surface; no auth built in (bind to localhost by default, put a
    real authenticating proxy in front for anything beyond a trusted network).
    See docs/INTEGRATION.md.

    qresp serve --port 8000
    curl -X POST localhost:8000/runs -H 'Content-Type: application/json' \\
      -d '{"questionnaire": "fixtures/eval/questionnaire_eval.xlsx", "provider": "stub"}'
    curl localhost:8000/runs/<id>
    curl localhost:8000/runs/<id>/results
    """
    import uvicorn

    from src.service import app

    click.echo(f"Serving the integration API on http://{host}:{port} — see docs/INTEGRATION.md.")
    uvicorn.run(app, host=host, port=port)



@cli.command()
@click.option("--yes", is_flag=True, help="Skip the confirmation prompt (non-interactive use).")
def purge(yes: bool):
    """Delete this workspace's store: SQLite store.db and the Chroma index (pack 3, C11).

    Retention answer a security team will ask for: how do we purge? This deletes
    the resolved data directory's store artifacts — the chunk text, every drafted
    answer, the audit trail — for the current workspace (qresp --workspace acme
    purge) or the default store (qresp purge). Run artifacts under the same data
    directory are NOT deleted; pass the directory to your normal file deletion
    process for a complete purge. Requires explicit confirmation unless --yes.
    """
    import shutil

    from src.data_dir import data_dir

    target = data_dir()
    paths = [target / "store.db", target / "chroma"]
    existing = [p for p in paths if p.exists()]
    if not existing:
        click.echo(f"Nothing to purge at {target} (no store.db or chroma/).")
        return
    if not yes:
        click.confirm(
            f"Delete the store at {target}? This removes {', '.join(p.name for p in existing)} "
            "— chunk text, answers, and the audit trail. This cannot be undone.",
            abort=True,
        )
    for p in existing:
        if p.is_dir():
            shutil.rmtree(p)
        else:
            p.unlink()
    click.echo(f"Purged {', '.join(str(p) for p in existing)}.")


def _aggregate_sub_results(sub_results):
    """Combine per-sub-question results into ONE row-level result (B3).

    The compound-question loop used to process every sub-question but let the LAST
    one win the cell while inserting one answers/audit row per sub-question — the
    review UI's LEFT JOIN fanned out and the progress bar double-counted. Contract:
    answer each part, then combine into exactly one row-level result before any
    write. Answer texts join with their sub-question as a lead-in (single-part
    rows keep the part's verbatim answer, so the pass-through is bit-identical to
    the pre-loop behaviour), cited chunks/sentences are unioned, confidence is the
    WEAKEST across parts (a row is only as trustworthy as its worst-supported
    part), polarity is "partial" on any disagreement, and the per-part detail rides
    in the row's structured log and answers row. Exactly one cell, one answers row,
    one audit entry, one count per sheet row.
    """
    parts = []
    cited_ids = set()
    cited_sentences: list[str] = []
    sources: list[str] = []
    all_evidence: list[RetrievedChunk] = []
    polarities = set()
    confidences = []
    vocab_set = set()
    sub_detail = []
    in_tok = out_tok = cache_read = cache_creation = entail_in = entail_out = 0
    for sub_question, result, evidence in sub_results:
        all_evidence.extend(evidence)
        sources.extend(f"{c.source_filename} ({c.heading_path or 'no heading'}, {c.loc_ref})" for c in evidence)
        in_tok += result.input_tokens
        out_tok += result.output_tokens
        cache_read += result.cache_read_input_tokens
        cache_creation += result.cache_creation_input_tokens
        entail_in += result.entailment_input_tokens
        entail_out += result.entailment_output_tokens
        polarities.add(result.polarity)
        cited_ids.update(result.cited_chunk_ids)
        cited_sentences.extend(result.cited_sentences)
        if result.vocab_selection is not None:
            vocab_set.add(result.vocab_selection)
        if result.status == AnswerStatus.NOT_FOUND:
            confidences.append("none")
            parts.append(f"{sub_question}: (no supporting evidence found)")
        elif result.status == AnswerStatus.ANSWERED:
            confidences.append(result.confidence or "low")
            parts.append(f"{sub_question}: {result.answer}")
        else:
            confidences.append("error")
            parts.append(f"{sub_question}: (processing error)")
        sub_detail.append(
            {
                "text": sub_question,
                "confidence": confidences[-1],
                "answer": result.answer,
                "polarity": result.polarity,
            }
        )
    weakest = (
        "none" if "none" in confidences else ("low" if "low" in confidences else "high" if confidences else "none")
    )
    non_null_polarities = {p for p in polarities if p is not None}
    if len(non_null_polarities) > 1:
        polarity = "partial"
    elif len(non_null_polarities) == 1:
        polarity = next(iter(non_null_polarities))
    else:
        polarity = None
    # Single-part rows (the pass-through) keep the part's verbatim answer and
    # confidence exactly as before; the lead-in join is only for multi-part rows.
    if len(sub_results) == 1:
        answer_text = sub_results[0][1].answer
    else:
        answer_text = chr(10).join(chr(10) + p for p in parts) if parts else ""
    return {
        "answer_text": answer_text,
        "final_confidence": weakest,
        "self_confidence": weakest if weakest != "none" else None,
        "polarity": polarity if weakest != "none" else None,
        "vocab_selection": next(iter(vocab_set)) if len(vocab_set) == 1 else None,
        "cited_chunk_ids": sorted(cited_ids),
        "cited_sentences": cited_sentences,
        "sources": sources,
        "evidence": all_evidence,
        "sub_question_text": chr(10).join(s for s, _, _ in sub_results),
        "sub_questions": sub_detail,
        "input_tokens": in_tok,
        "output_tokens": out_tok,
        "cache_read_input_tokens": cache_read,
        "cache_creation_input_tokens": cache_creation,
        "entailment_input_tokens": entail_in,
        "entailment_output_tokens": entail_out,
    }


def _add_stub_banner(ws, last_col: int) -> None:
    banner_row = ws.max_row + 1
    ws.merge_cells(start_row=banner_row, start_column=1, end_row=banner_row, end_column=last_col)
    cell = ws.cell(row=banner_row, column=1, value=STUB_BANNER_TEXT)
    cell.fill = STUB_BANNER_FILL
    cell.font = STUB_BANNER_FONT
    cell.alignment = Alignment(horizontal="center")


@cli.command()
@click.option("--questionnaire", type=click.Path(exists=True, dir_okay=False, path_type=Path), required=True)
@click.option("--output", type=click.Path(path_type=Path), required=True)
@click.option(
    "--limit",
    type=int,
    default=5,
    show_default=True,
    help="Max question rows to process this run. Use 0 for no limit (process every detected question row).",
)
@click.option(
    "--only-row",
    type=int,
    default=None,
    help="Process only this single sheet row (overrides --limit); useful for targeted checks.",
)
@click.option(
    "--sheet",
    type=str,
    default=None,
    help="Process only this worksheet by name (default: every sheet with a detectable question/answer header).",
)
@click.option(
    "--map",
    "map_override",
    type=str,
    default=None,
    help="Override column detection entirely: --map question=C,answer=E,vocab=D (letters or numbers). "
    "Use after 'qresp inspect' when detection guesses wrong on a real file.",
)
@click.option("--provider", type=click.Choice(["anthropic", "stub", "local"]), default="anthropic", show_default=True)
@click.option(
    "--stub-fail-row",
    type=int,
    default=None,
    help="With --provider stub, make that row raise, to exercise per-row error isolation.",
)
@click.option(
    "--dry-run",
    is_flag=True,
    help="Estimate what a run will cost before starting it: column detection, question reading, and retrieval for every selected row, then real (local) token counts and an estimated price. Makes zero API calls, writes no workbook, and creates no run row.",
)
@click.option(
    "--config",
    type=click.Path(exists=True, dir_okay=False, path_type=Path),
    default=None,
    help="TOML file of tuning knobs (see src/config.py). Precedence: CLI flags > QRESP_* env vars > this file > defaults.",
)
@click.option(
    "--top-k",
    type=int,
    default=None,
    help="Retrieved chunks per question (highest-precedence override of the config's top_k).",
)
@click.option(
    "--exact",
    is_flag=True,
    help="With --dry-run: use the free count_tokens API for true input-token counts instead of the local tokenizer's 1.4-1.9x undercount band (requires ANTHROPIC_API_KEY; generates nothing).",
)
@click.option(
    "--verbose",
    is_flag=True,
    help="Log per-row detail (including tracebacks) to stderr; also always written to the structured <output>.log.jsonl file.",
)
@click.option(
    "--quiet",
    is_flag=True,
    help="Suppress INFO progress logging on stderr (errors only); the structured log file is unaffected.",
)
def answer(
    questionnaire: Path,
    output: Path,
    limit: int,
    only_row: int | None,
    sheet: str | None,
    map_override: str | None,
    provider: str,
    stub_fail_row: int | None,
    dry_run: bool,
    config: Path | None,
    top_k: int | None,
    exact: bool,
    verbose: bool,
    quiet: bool,
):
    from src.config import load_config

    if verbose and quiet:
        raise click.ClickException("--verbose and --quiet are mutually exclusive.")

    # Guard against clobbering the customer's source workbook: --output resolving
    # to the same file as --questionnaire (directly or via a symlink — resolve()
    # follows symlinks) would overwrite it in place, unrecoverably. Runs BEFORE
    # _setup_logging, which would otherwise create the log file next to the
    # rejected path.
    if questionnaire.resolve() == output.resolve():
        raise click.ClickException(
            f"--output {output} is the same file as --questionnaire {questionnaire} — "
            f"refusing to overwrite the source workbook in place. Choose a different --output."
        )
    # The output parent must exist BEFORE logging setup: the structured log file
    # opens next to the workbook, and a first run into a fresh directory (e.g.
    # --output reports/filled.xlsx with no reports/ yet) used to crash here,
    # before the try/finally that would have created it.
    output.parent.mkdir(parents=True, exist_ok=True)
    _setup_logging(output, verbose=verbose, quiet=quiet)

    cli_overrides = {}
    if top_k is not None:
        cli_overrides["top_k"] = top_k
    cfg = load_config(config_file=config, cli_overrides=cli_overrides)

    answerer: Answerer
    if provider == "anthropic" and not dry_run:
        if not os.environ.get("ANTHROPIC_API_KEY"):
            raise click.ClickException(
                "ANTHROPIC_API_KEY not set — export it before running with --provider anthropic (every row would fail identically). Use --provider stub if you want to test without a key."
            )
        answerer = AnthropicAnswerer(config=cfg)
    elif provider == "local" and not dry_run:
        # C7: fully on-premise generation via an OpenAI-compatible endpoint
        # (Ollama, vLLM, llama.cpp server). No API key, nothing leaves the host;
        # the citation grounding and entailment checks run identically (see
        # src/answer/local.py). Configure with QRESP_LOCAL_BASE_URL /
        # QRESP_LOCAL_MODEL or a config file.
        from src.answer.local import LocalAnswerer, LocalConfig

        answerer = LocalAnswerer(
            LocalConfig(base_url=cfg.local_base_url, model=cfg.local_model),
            weak_match_distance=cfg.weak_match_distance,
            entailment_check=cfg.entailment_check,
            entailment_model=cfg.local_model,
        )
        click.echo(f"Using local model {cfg.local_model} at {cfg.local_base_url} — nothing leaves this machine.")
    else:
        answerer = StubAnswerer(fail_row=stub_fail_row, weak_match_distance=cfg.weak_match_distance)

    workbook = openpyxl.load_workbook(questionnaire)
    from src.questionnaire.parse_xlsx import _parse_column_override, iter_question_sheets

    column_override = _parse_column_override(map_override)
    if sheet is not None:
        if sheet not in workbook.sheetnames:
            raise click.ClickException(
                f"Workbook has no sheet named {sheet!r} — sheets: {', '.join(workbook.sheetnames)}"
            )
        ws = workbook[sheet]
        column_map = detect_columns(ws, column_override=column_override)
        sheet_questions = [(sheet, ws, column_map, read_questions(ws, column_map))]
    else:
        sheet_questions = [
            (name, ws, cm, read_questions(ws, cm))
            for name, ws, cm in iter_question_sheets(workbook, column_override=column_override)
        ]
        if not sheet_questions:
            raise click.ClickException(
                "No sheet in the workbook has a detectable question/answer header — run "
                "'qresp inspect --questionnaire FILE' to see what was detected, and use "
                "--map question=C,answer=E,vocab=D if a real file's headers are unconventional."
            )
    total_detected = sum(len(qs) for _, _, _, qs in sheet_questions)
    if only_row is not None:
        selected = [(s, ws, cm, q) for s, ws, cm, qs in sheet_questions for q in qs if q.row_index == only_row]
        if not selected:
            raise click.ClickException(f"Row {only_row} is not a detected question row.")
    else:
        flat = [(s, ws, cm, q) for s, ws, cm, qs in sheet_questions for q in qs]
        selected = flat if limit == 0 else flat[:limit]
    sheets_note = f" across {len(sheet_questions)} sheet(s)" if len(sheet_questions) > 1 else ""
    click.echo(f"Answering {len(selected)} row(s) ({total_detected} detected){sheets_note} with provider={provider}.")
    log.info(
        "answer run starting: %d row(s) (%d detected across %d sheets), provider=%s, dry_run=%s",
        len(selected), total_detected, len(sheet_questions), provider, dry_run,
    )

    if dry_run:
        _dry_run_cost_estimate(selected, total_detected, cfg, exact=exact)
        return

    if provider == "stub":
        for _, ws, column_map, _ in sheet_questions:
            _add_stub_banner(ws, last_col=max(column_map.question_col, column_map.answer_col, column_map.vocab_col or 0))

    conn = db.connect()
    vector_store = VectorStore(model_name=cfg.embedding_model)
    reranker = None
    if cfg.reranker:
        from src.retrieval.reranker import CrossEncoderReranker

        # Downloads BAAI/bge-reranker-base on first use (~1.1 GB, local, no API
        # cost) — see the README setup section. One instance for the whole run.
        click.echo("Loading reranker model (first use downloads ~1.1 GB) ...")
        reranker = CrossEncoderReranker()
    searcher = HybridSearcher(
        conn,
        vector_store,
        vector_weight=cfg.vector_weight,
        rrf_k=cfg.rrf_k,
        candidate_pool=cfg.candidate_pool,
        reranker=reranker,
    )
    run_config = _current_run_config(cfg)
    click.echo(f"Config: {_config_fingerprint(run_config)}")
    run_id = db.start_questionnaire_run(conn, str(questionnaire), str(output), run_config=run_config)

    output.parent.mkdir(parents=True, exist_ok=True)
    jsonl_path = output.with_suffix(".jsonl")

    counts = {"high": 0, "low": 0, "none": 0, "error": 0}
    total_input_tokens = 0
    total_output_tokens = 0
    total_cache_read_tokens = 0
    total_cache_creation_tokens = 0
    total_entailment_input_tokens = 0
    total_entailment_output_tokens = 0
    consecutive_errors = 0
    caught_exc = (
        None  # set on the error path so the structured log can carry the traceback after the except block exits
    )

    try:
        with open(jsonl_path, "a") as jsonl_file:
            for i, (sheet_name, ws, column_map, q) in enumerate(selected, start=1):
                start = time.monotonic()
                # B3 aggregation contract: split_question is a pass-through today,
                # but the loop must be written for the day it isn't. Each
                # sub-question gets a full retrieve/generate cycle, then the parts
                # are combined into ONE row-level result before any write — one
                # cell, one answers row, one audit entry, one count per sheet row.
                sub_results = []
                error_message: str | None = None
                # C4: answer library — prior human-approved answers surfaced as labelled
                # candidates for the generator (never as retrieval evidence). Looked up
                # once per row before the sub-question loop; freshness-gated inside
                # find_candidates (stale source docs exclude an entry).
                prior_answers: list[dict] = []
                if cfg.answer_library:
                    from src.answer.library import find_candidates

                    prior_answers = find_candidates(
                        conn, q.question_text, vector_store, threshold=cfg.library_semantic_threshold
                    )
                    if prior_answers:
                        log.debug(
                            "answer library: %d candidate(s) surfaced for row %s",
                            len(prior_answers),
                            q.row_index,
                        )
                try:
                    for sub_question in split_question(q.question_text):
                        # B1: evidence must be bound BEFORE search — if search
                        # raises on the first row, the error handler below would
                        # die with NameError while logging the failure; on any
                        # later row it would hold stale chunks from the previous
                        # row and record the wrong retrieval results.
                        evidence = []
                        evidence = searcher.search(sub_question, top_k=cfg.top_k)
                        result = answerer.answer_question(
                            sub_question, evidence, column_map.vocab_values, row_index=q.row_index,
                            prior_answers=prior_answers or None,
                        )
                        sub_results.append((sub_question, result, evidence))
                        total_input_tokens += result.input_tokens
                        total_output_tokens += result.output_tokens
                        total_cache_read_tokens += result.cache_read_input_tokens
                        total_cache_creation_tokens += result.cache_creation_input_tokens
                        total_entailment_input_tokens += result.entailment_input_tokens
                        total_entailment_output_tokens += result.entailment_output_tokens
                except FATAL_ERRORS as exc:
                    # Wrong key, wrong model, or a schema-rejecting request: the same
                    # failure for every remaining row. Abort now (finally saves
                    # everything processed so far) instead of failing 400 rows one at
                    # a time after a full retry ladder each.
                    click.echo(f"  row {q.row_index}: FATAL — {type(exc).__name__}: {exc}")
                    raise click.ClickException(
                        f"Run aborted on row {q.row_index}: {type(exc).__name__} — this error will "
                        f"repeat for every row (check the API key, model name, and request schema). "
                        f"Everything processed so far is saved to {output}."
                    ) from exc
                except Exception as exc:
                    consecutive_errors += 1
                    caught_exc = exc
                    # A row that burned API calls and then raised used to report zero
                    # tokens. generate_answer attaches the real usage to the exception
                    # (see generate.record_usage), so the run summary reflects money
                    # actually spent even on failed rows.
                    row_usage = getattr(exc, "_row_usage", None)
                    if row_usage:
                        total_input_tokens += row_usage["input_tokens"]
                        total_output_tokens += row_usage["output_tokens"]
                        total_cache_read_tokens += row_usage["cache_read_input_tokens"]
                        total_cache_creation_tokens += row_usage["cache_creation_input_tokens"]
                    if consecutive_errors >= CONSECUTIVE_ERROR_LIMIT:
                        click.echo(f"  row {q.row_index}: ERROR — {exc}")
                        raise click.ClickException(
                            f"Run aborted after {CONSECUTIVE_ERROR_LIMIT} consecutive per-row errors "
                            f"(circuit breaker) — the failure is systemic, not per-row. Last error: {exc}. "
                            f"Everything processed so far is saved to {output}."
                        ) from exc
                    final_confidence = "error"
                    cited_chunk_ids = []
                    cited_sentences = []
                    sources = []
                    evidence = []
                    library_state = None
                    library_provenance = None
                    answer_text = None
                    vocab_selection = None
                    self_confidence = None
                    polarity = None
                    sub_question_text = q.question_text
                    in_tokens = (row_usage or {}).get("input_tokens")
                    out_tokens = (row_usage or {}).get("output_tokens")
                    entail_in = entail_out = 0
                    sub_detail = []
                    error_message = str(exc)
                    click.echo(f"  row {q.row_index}: ERROR — {error_message}")
                else:
                    consecutive_errors = 0
                    agg = _aggregate_sub_results(sub_results)
                    final_confidence = agg["final_confidence"]
                    answer_text = agg["answer_text"]
                    vocab_selection = agg["vocab_selection"]
                    self_confidence = agg["self_confidence"]
                    polarity = agg["polarity"]
                    cited_chunk_ids = agg["cited_chunk_ids"]
                    cited_sentences = agg["cited_sentences"]
                    sources = agg["sources"]
                    evidence = agg["evidence"]
                    sub_question_text = agg["sub_question_text"]
                    # C4: mark rows that drew on the library. A row counts as "used"
                    # when a fresh candidate was surfaced AND the generated answer
                    # materially reuses it; "surfaced" when a candidate existed but
                    # the draft went its own way. Either way the reviewer sees the
                    # provenance, and the citation/entailment checks have already
                    # run against the ORIGINAL evidence.
                    library_state = None
                    library_provenance = None
                    if prior_answers and final_confidence in ("high", "low"):
                        from src.answer.library import answer_uses_prior

                        best = prior_answers[0]
                        used = answer_uses_prior(answer_text or "", best)
                        library_state = "used" if used else "surfaced"
                        library_provenance = (
                            f"run {best['run_id']} row {best['row_index']} "
                            f"{best['human_action']} at {best['reviewed_at']}"
                        )
                    in_tokens = agg["input_tokens"]
                    out_tokens = agg["output_tokens"]
                    entail_in = agg["entailment_input_tokens"]
                    entail_out = agg["entailment_output_tokens"]
                    sub_detail = agg["sub_questions"]
                    error_message = None

                elapsed = time.monotonic() - start

                row_data = {
                    "row_index": q.row_index,
                    "final_confidence": final_confidence,
                    "polarity": polarity,
                    "provider": provider,
                    "elapsed_seconds": round(elapsed, 3),
                    "retrieval": [
                        {
                            "embedding_id": c.embedding_id,
                            "combined_score": round(c.combined_score, 4),
                            "distance": c.vector_distance,
                        }
                        for c in evidence
                    ],
                    "sub_questions": sub_detail,
                    "input_tokens": in_tokens,
                    "output_tokens": out_tokens,
                    "entailment_input_tokens": entail_in,
                    "entailment_output_tokens": entail_out,
                    "library_state": library_state,
                    "library_provenance": library_provenance,
                    "error": error_message,
                }
                if error_message is None:
                    log.debug(
                        "row %s: %s (%.1fs)", q.row_index, final_confidence, elapsed, extra={"row_data": row_data}
                    )
                else:
                    log.error(
                        "row %s: ERROR — %s",
                        q.row_index,
                        error_message,
                        exc_info=(type(caught_exc), caught_exc, caught_exc.__traceback__) if caught_exc else True,
                        extra={"row_data": row_data},
                    )
                caught_exc = None

                write_answer(
                    ws, q.row_index, column_map, answer_text or "", vocab_selection, final_confidence,
                    library_state=library_state, library_provenance=library_provenance,
                )
                db.record_answer(
                    conn,
                    run_id,
                    q.row_index,
                    q.question_text,
                    sub_question_text,
                    answer_text,
                    vocab_selection,
                    self_confidence,
                    final_confidence,
                    cited_chunk_ids,
                    polarity=polarity,
                    cited_sentences=cited_sentences,
                    sheet_name=sheet_name,
                    library_candidate=(
                        {
                            "state": library_state,
                            "provenance": library_provenance,
                            "candidates": [
                                {
                                    "run_id": c["run_id"],
                                    "row_index": c["row_index"],
                                    "human_action": c["human_action"],
                                    "reviewed_at": c["reviewed_at"],
                                    "similarity": c.get("similarity"),
                                }
                                for c in prior_answers
                            ],
                        }
                        if prior_answers
                        else None
                    ),
                )
                db.record_audit_entry(conn, run_id, q.row_index, sources, final_confidence, provider=provider)

                jsonl_file.write(
                    json.dumps(
                        {
                            # run_id + timestamp on every record: the sidecar opens in append
                            # mode, and without them two runs to the same --output interleave
                            # into one file with no way to separate them (P20).
                            "run_id": run_id,
                            "ts": datetime.now(UTC).isoformat(),
                            "sheet": sheet_name,
                            "row_index": q.row_index,
                            "question_text": q.question_text,
                            "final_confidence": final_confidence,
                            "polarity": polarity,
                            "provider": provider,
                            "library_state": library_state,
                            "answer": answer_text,
                            "error": error_message,
                            "elapsed_seconds": round(elapsed, 2),
                        }
                    )
                    + "\n"
                )
                jsonl_file.flush()

                counts[final_confidence] += 1
                if error_message is None:
                    click.echo(f"  row {q.row_index}: {final_confidence} ({elapsed:.1f}s)")
                if i % SAVE_EVERY_N_ROWS == 0:
                    workbook.save(output)
    finally:
        # Guarantees the xlsx always reflects everything processed so far, even on an
        # unhandled exception or Ctrl-C — the periodic save above is just a progress
        # convenience, not what correctness depends on.
        workbook.save(output)

    n_answered = counts["high"] + counts["low"]
    click.echo(
        f"Wrote {output} (+ {jsonl_path.name}). "
        f"answered={n_answered} flagged_low={counts['low']} not_found={counts['none']} error={counts['error']}"
    )
    log.info(
        "answer run finished: answered=%d flagged_low=%d not_found=%d error=%d",
        n_answered,
        counts["low"],
        counts["none"],
        counts["error"],
    )
    if provider == "anthropic" and (total_input_tokens or total_output_tokens):
        uncached_input = total_input_tokens - total_cache_read_tokens
        # B2: per-question averages only make sense when something was answered —
        # a run where every row came back NOT_FOUND still spends real tokens on
        # every row, and that is precisely the run whose cost you want to see, so
        # the totals print regardless and only the averages are guarded.
        avg_line = (
            f"(avg {total_input_tokens // n_answered} in / {total_output_tokens // n_answered} out per answered question)"
            if n_answered > 0
            else "(no answered questions)"
        )
        click.echo(f"Tokens: {total_input_tokens} in / {total_output_tokens} out {avg_line}")
        click.echo(
            f"  of which {uncached_input} uncached in, {total_cache_read_tokens} cache-read in, "
            f"{total_cache_creation_tokens} cache-created in"
        )
        if total_entailment_input_tokens or total_entailment_output_tokens:
            entail_cost = _estimate_cost(
                total_entailment_input_tokens,
                total_entailment_output_tokens,
                cfg.input_price_per_mtok,
                cfg.output_price_per_mtok,
            )
            click.echo(
                f"  entailment check (A1): {total_entailment_input_tokens} in / "
                f"{total_entailment_output_tokens} out (est. ${_format_cost(entail_cost)})"
            )
        total_cost = _estimate_cost(
            total_input_tokens, total_output_tokens, cfg.input_price_per_mtok, cfg.output_price_per_mtok
        )
        click.echo(f"  estimated cost: ${_format_cost(total_cost)}")


# Measured output-token average per ANSWERED question from the deterministic
# 20-question baseline eval runs (55755 input / 7927 output over 7 answered rows;
# 1132 out per answered row). Used by --dry-run to estimate output spend; the
# answered-only average overestimates for a run with many NOT_FOUND rows, which is
# the conservative direction for cost planning.
OUTPUT_TOKENS_PER_ANSWERED_QUESTION = 1132


def _dry_run_cost_estimate(selected: list, total_detected: int, cfg, exact: bool = False) -> None:
    """Everything up to the API call, then stop: column detection and question
    reading already happened; here we run retrieval for every selected row, count
    input tokens — system prompt plus the exact user message generate_answer would
    build — add an output-token estimate from observed averages, and print an
    estimated cost. No workbook written, no run row created in questionnaire_runs.

    The estimate must describe the run that would actually happen, so the searcher
    and store are built IDENTICALLY to the real path from the resolved Config —
    top_k, embedding model, fusion constants, and the reranker (B4).

    Two counting modes: local (default) uses src.answer.tokenize, which ships the
    real Claude BPE locally and is documented to undercount the live model by
    1.4-1.9x — so the cost prints as a RANGE over that band, and the undercount is
    in the dangerous direction for budgeting. exact uses the count_tokens API for
    true counts: it IS an API call, but it is free and generates nothing — the
    dry-run's zero-spend rule is about money, and --exact makes the trade
    explicit."""
    from src.answer.generate import SYSTEM_PROMPT, _build_user_message
    from src.answer.split_questions import split_question
    from src.answer.tokenize import count_tokens
    from src.retrieval.hybrid_search import HybridSearcher
    from src.store import db
    from src.store.vectorstore import VectorStore

    conn = db.connect()
    vector_store = VectorStore(model_name=cfg.embedding_model)
    reranker = None
    if cfg.reranker:
        from src.retrieval.reranker import CrossEncoderReranker

        reranker = CrossEncoderReranker()
    searcher = HybridSearcher(
        conn,
        vector_store,
        vector_weight=cfg.vector_weight,
        rrf_k=cfg.rrf_k,
        candidate_pool=cfg.candidate_pool,
        reranker=reranker,
    )

    system_tokens = count_tokens(SYSTEM_PROMPT)
    total_input = 0
    total_retrieved = 0
    count_method = (
        "local Claude tokenizer (claude-v1 BPE; undercounts the live model by ~1.4-1.9x — see src/answer/tokenize.py)"
    )
    if exact:
        # count_tokens is a real API call but free and generates nothing; the
        # dry-run's zero-spend rule is about money, and --exact makes the trade
        # explicit (B5).
        import os

        if not os.environ.get("ANTHROPIC_API_KEY"):
            raise click.ClickException(
                "--dry-run --exact needs ANTHROPIC_API_KEY (it calls the free count_tokens API). "
                "Without a key, use plain --dry-run with the local tokenizer."
            )
        import anthropic

        client = anthropic.Anthropic(max_retries=0, timeout=30.0)
        count_method = "count_tokens API (exact)"
        for sheet_name, ws, column_map, q in selected:
            sub_question = split_question(q.question_text)[0]
            evidence = searcher.search(sub_question, top_k=cfg.top_k)
            user_content = _build_user_message(sub_question, evidence, column_map.vocab_values)
            resp = client.messages.count_tokens(
                model=cfg.model,
                system=[{"type": "text", "text": SYSTEM_PROMPT}],
                messages=[{"role": "user", "content": user_content}],
            )
            total_input += resp.input_tokens
            total_retrieved += len(evidence)
    else:
        for sheet_name, ws, column_map, q in selected:
            sub_question = split_question(q.question_text)[0]
            evidence = searcher.search(sub_question, top_k=cfg.top_k)
            user_content = _build_user_message(sub_question, evidence, column_map.vocab_values)
            total_input += system_tokens + count_tokens(user_content)
            total_retrieved += len(evidence)

    output_estimate = OUTPUT_TOKENS_PER_ANSWERED_QUESTION * len(selected)

    click.echo(
        "Dry run — " + ("no API calls, " if not exact else "one free count_tokens call per row, ") + "nothing written:"
    )
    click.echo("  rows detected: " + str(total_detected) + "   rows selected: " + str(len(selected)))
    click.echo(
        "  estimated input tokens: "
        + str(total_input)
        + " (system prompt "
        + str(system_tokens if not exact else 0)
        + " tokens per row + retrieved chunks + question, counted with the "
        + count_method
        + ")"
    )
    click.echo(
        "  estimated output tokens: "
        + str(output_estimate)
        + " (observed avg "
        + str(OUTPUT_TOKENS_PER_ANSWERED_QUESTION)
        + " per answered question — output is never counted, only input is)"
    )
    if exact:
        cost = _estimate_cost(total_input, output_estimate, cfg.input_price_per_mtok, cfg.output_price_per_mtok)
        click.echo("  estimated cost: $" + _format_cost(cost) + " (exact input count; output still an estimate)")
    else:
        # The local tokenizer is documented to undercount the live model by
        # 1.4-1.9x; budget for the worst case by reporting a RANGE over that band
        # (B5) instead of a false-precision point estimate.
        low = _estimate_cost(
            int(total_input * 1.4), output_estimate, cfg.input_price_per_mtok, cfg.output_price_per_mtok
        )
        high = _estimate_cost(
            int(total_input * 1.9), output_estimate, cfg.input_price_per_mtok, cfg.output_price_per_mtok
        )
        click.echo(
            "  estimated cost: $"
            + _format_cost(low)
            + "-$"
            + _format_cost(high)
            + " (input-tokenizer undercount band 1.4-1.9x; --exact for a true count)"
        )


def _current_run_config(cfg) -> dict:
    """Snapshot of the resolved configuration that produced this run, for the
    run_config column and the start-of-run fingerprint.

    Serializes the P18 Config (all answer-affecting values: model, token limit,
    confidence threshold, fusion constants, pool sizes, chunk bounds, embedding
    model) plus the git revision and a dirty-tree flag, so a past run in
    out/store.db can always be tied back to the exact code+config that produced it."""
    import subprocess

    repo_root = Path(__file__).resolve().parent.parent
    git_sha = None
    git_dirty = None
    try:
        git_sha = subprocess.check_output(
            ["git", "rev-parse", "HEAD"], cwd=repo_root, stderr=subprocess.DEVNULL, text=True
        ).strip()
        git_dirty = bool(subprocess.check_output(["git", "status", "--porcelain"], cwd=repo_root, text=True).strip())
    except Exception as exc:  # noqa: BLE001 — not in a git repo or git unavailable: record the gap, don't crash
        log.debug("git revision unavailable: %s", exc)

    snapshot = cfg.as_dict()
    snapshot["git_sha"] = git_sha
    snapshot["git_dirty"] = git_dirty
    return snapshot


def _config_fingerprint(cfg: dict) -> str:
    """One-line fingerprint printed at the start of every answer run, so the
    terminal shows the config before any money is spent on it."""
    git = (cfg.get("git_sha") or "no-git") + ("(dirty)" if cfg.get("git_dirty") else "")
    return (
        f"model={cfg['model']} max_tokens={cfg['max_tokens']} "
        f"weak_match={cfg['weak_match_distance']} vector_weight={cfg['vector_weight']} "
        f"rrf_k={cfg['rrf_k']} pool={cfg['candidate_pool']} top_k={cfg['top_k']} "
        f"chunks={cfg['max_chunk_chars']}/{cfg['min_chunk_chars']}/{cfg['overlap_sentences']} "
        f"embed={cfg['embedding_model']} git={git}"
    )


def _estimate_cost(
    input_tokens: int,
    output_tokens: int,
    input_price_per_mtok: float = 3.0,
    output_price_per_mtok: float = 15.0,
) -> float:
    """Rough dollar estimate for the tokens reported this run.

    Prices come from the Config rate card (B5) so a price change is a config
    change, not a code edit; the defaults mirror the old source constants. The
    true numbers depend on the current model's published rate card and any
    caching discounts, so this is an order-of-magnitude estimate for capacity
    planning, not a bill. Cache-read input tokens are billed at a fraction of
    fresh input tokens; the estimate conservatively prices everything at the
    fresh rate, which overestimates once prompt caching engages."""
    input_cost = input_tokens / 1_000_000 * input_price_per_mtok
    output_cost = output_tokens / 1_000_000 * output_price_per_mtok
    return input_cost + output_cost


def _format_cost(x: float) -> str:
    """Two significant figures — a cost estimate with four decimals reads as a
    precision the number does not have (B5)."""
    return f"{x:.2g}"


if __name__ == "__main__":
    cli()
